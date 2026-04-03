import os
os.chdir(os.path.dirname(os.path.abspath(__file__)))

import io
import hashlib
import json
import base64
import requests
import unicodedata
import streamlit as st

# ── INICIALIZAR SESSION STATE ─────────────────────────────────────
for _k in ["resultados_cart","informe_cart","resultados_enc","informe_enc"]:
    if _k not in st.session_state:
        st.session_state[_k] = None

import pandas as pd
import numpy as np
from sentence_transformers import SentenceTransformer
from sklearn.cluster import KMeans
from sklearn.metrics import silhouette_score
from sklearn.metrics.pairwise import cosine_similarity
import plotly.express as px
from umap import UMAP
from mistralai import Mistral

# openpyxl para Excel con fórmulas
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Análisis Semántico", layout="wide")
st.title("🔍 Análisis Semántico de Encuestas y Cartografía Social")

client = Mistral(api_key=st.secrets["MISTRAL_API_KEY"])

@st.cache_resource
def cargar_modelo():
    return SentenceTransformer("paraphrase-multilingual-mpnet-base-v2")

modelo = cargar_modelo()

# ════════════════════════════════════════════════════════════════
# SISTEMA DE CACHÉ — GITHUB PERSISTENCIA
# ════════════════════════════════════════════════════════════════
GITHUB_REPO  = "rezpemario-cell/analisis-semantico"
CACHE_FOLDER = "cache"

def md5_archivo(archivo):
    """Calcula hash MD5 del contenido del archivo. Siempre igual para el mismo archivo."""
    archivo.seek(0)
    contenido = archivo.read()
    archivo.seek(0)
    return hashlib.md5(contenido).hexdigest()

def leer_cache_github(hash_md5):
    """Busca cache/{hash_md5}.json en GitHub. Retorna dict o None."""
    try:
        token = st.secrets.get("GITHUB_TOKEN", "")
        if not token:
            return None
        url = (f"https://api.github.com/repos/{GITHUB_REPO}"
               f"/contents/{CACHE_FOLDER}/{hash_md5}.json")
        headers = {"Authorization": f"token {token}",
                   "Accept": "application/vnd.github.v3+json"}
        r = requests.get(url, headers=headers, timeout=10)
        if r.status_code == 200:
            contenido_b64 = r.json()["content"]
            contenido = base64.b64decode(contenido_b64).decode("utf-8")
            return json.loads(contenido)
    except Exception:
        pass
    return None

def escribir_cache_github(hash_md5, datos):
    """Guarda datos como JSON en GitHub. Actualiza si ya existe. Retorna True/False."""
    try:
        token = st.secrets.get("GITHUB_TOKEN", "")
        if not token:
            return False
        url = (f"https://api.github.com/repos/{GITHUB_REPO}"
               f"/contents/{CACHE_FOLDER}/{hash_md5}.json")
        headers = {"Authorization": f"token {token}",
                   "Accept": "application/vnd.github.v3+json"}
        contenido_str = json.dumps(datos, ensure_ascii=False, default=str)
        contenido_b64 = base64.b64encode(contenido_str.encode("utf-8")).decode("utf-8")
        r_get = requests.get(url, headers=headers, timeout=10)
        payload = {"message": f"cache: {hash_md5[:8]}", "content": contenido_b64}
        if r_get.status_code == 200:
            payload["sha"] = r_get.json()["sha"]
        r_put = requests.put(url, headers=headers, json=payload, timeout=15)
        return r_put.status_code in [200, 201]
    except Exception:
        return False

def df_a_cache(df):
    """Convierte DataFrame a lista de dicts JSON-serializable."""
    records = []
    for _, row in df.iterrows():
        rec = {}
        for col in df.columns:
            val = row[col]
            if isinstance(val, list):
                rec[col] = val
            elif not isinstance(val, (list, dict)) and pd.isna(val):
                rec[col] = None
            else:
                rec[col] = val
        records.append(rec)
    return records

def cache_a_df(records):
    """Reconstruye DataFrame desde lista de dicts."""
    return pd.DataFrame(records)

# ════════════════════════════════════════════════════════════════
# DICCIONARIO DE TEMAS (módulo global)
# ════════════════════════════════════════════════════════════════
KEYWORDS_TEMAS = {
    'Agua y medio ambiente': [
        'agua','acueducto','hídrico','fuentes','nacimiento','reforestar','árbol',
        'medio ambiente','ambiental','cuenca','microcuenca','monitoreo','conservación',
        'reforestación','siembra','sembrar','sequía','recursos naturales','natural',
        'recurso','flora','fauna','ecosistema','rio','quebrada'],
    'Liderazgo y organización comunitaria': [
        'liderazgo','líder','jac','junta','organización','participación','capacitar',
        'fortalecer','fortalecimiento','vocería','empoderamiento','gestión',
        'comunidad','comunidades','unido','unir','unidos','unida','unidas',
        'participando','participar','acercan','acercar','colectiv','trabajar',
        'vereda','motivar','motivan','vision','pensar','conceptos','proceso'],
    'Mujer y género': [
        'mujer','mujeres','femenino','género','ama de casa','voceras','rol',
        'madre','madres','cabeza de hogar'],
    'Educación y formación': [
        'educación','formación','capacitación','aprendizaje','estudiante',
        'jóvenes','universidad','técnica','sena','conocimiento','habilidades',
        'derechos','deberes','entendimiento','diálogo','dialogo','concientizar',
        'niños','voceros','escuela'],
    'Infraestructura': [
        'vía','vias','camino','caseta','placa','huella','construcción','obra',
        'cancha','tanque','sede','gimnasio','gimnacio','infraestructura',
        'infracstuturas','abras','obras','carriles','movilidad','urbano',
        'pública','publicas','instalacion'],
    'Agricultura y economía rural': [
        'café','caficultura','agricultura','campo','emprendimiento','productivo',
        'producción','semilla','cosecha','finca','pollos','horticultura',
        'cultivo','siembra','campo','calidad de vida'],
    'Continuidad y sostenibilidad': [
        'continuar','continúe','seguir','siga','sostenible','sostenibilidad',
        'durar','mantenerse','frutos','adelante','futuro','presente','avanzar',
        'resultados','espectativas','expectativas','luchar','luchando'],
    'Comunicación e información': [
        'comunicación','información','transparencia','redes sociales','difusión',
        'visibilizar','informar','derechos','deberes','entendimiento','diálogo',
        'preguntas','personales'],
    'Mejora y sugerencias': [
        'mejorar','mejora','sugerencia','falta','necesita','cumplir',
        'cumplimiento','prometieron','convocatoria','compromiso','prioridad',
        'sector','cumplan','proyectos que nos dijeron','mas compromiso'],
    'Percepción positiva empresa': [
        'empresa','confianza','apoyo','acompañamiento','agradecimiento',
        'colaboración','beneficio','bueno','positivo','experiencia','maravillos',
        'bien','muy bien','todo bien','todo ha sido','ha sido positiv',
        'llenando','expectativas','espectativas'],
    'Riesgo y preocupación cierre': [
        'triste','abandono','finalizar','cierre','solos','sin apoyo','dejar',
        'retiro','desamparados','negativo','negativa','retroceso','mal hecho',
        'poco serios','motivos','algo malo','algo no les gusto','estrategia',
        'ingresar','territorio','desicion','decisión','no podemos hacer nada'],
    'Alianzas institucionales': [
        'alianza','alcaldía','gobernación','articulación','público-privada',
        'entidad','institucional'],
    'Minería': [
        'minería','minero','mina','formalización','exploración'],
    'Nuevos proyectos propuestos': [
        'turismo','turístico','aves','avistamiento','fauna','flora','tiempo libre',
        'ludico','pollos','horticultura','deporte','fútbol','futbol','deportiv',
        'piscina','recreacion','recreación','aire libre','salud','actividad física',
        'escuela de futbol','nueva actividad','nuevo proyecto'],
    'Respuesta inespecífica': [
        'ninguno','nada','ninguna','no se','no sé','no me ocurre','no ocurre',
        'no haría','no hariamos','no podemos','no se me','desecsion','desicion',
        'decisión','así esta','asi esta','esta bien','está bien','estamos en',
        'solo participé','solo participe','socializacion','socialización',
        'diferentes necesidades','en diferentes','no estaria','no estaría'],
}

LINEAS_CANONICAS = [
    "Agua y territorio",
    "Desarrollo rural",
    "Educación y competitividad",
    "Infraestructura comunitaria",
    "Fortalecimiento comunitario"
]

# ════════════════════════════════════════════════════════════════
# FUNCIONES AUXILIARES GENERALES
# ════════════════════════════════════════════════════════════════
def encontrar_clusters_optimos(vectores, min_k=2, max_k=8):
    max_k = min(max_k, len(vectores) - 1)
    if max_k < 2:
        return 2
    scores = {}
    for k in range(2, max_k + 1):
        km = KMeans(n_clusters=k, random_state=42, n_init=10)
        labels = km.fit_predict(vectores)
        scores[k] = silhouette_score(vectores, labels)
    return max(scores, key=scores.get)

def etiquetar_grupos_ia(frases_por_grupo, contexto):
    etiquetas = {}
    for grupo, frases in frases_por_grupo.items():
        muestra = "\n".join(f"- {f}" for f in frases[:5])
        prompt = (f"Eres un experto en análisis social y organizacional.\n"
                  f"Analiza estas frases de un grupo semántico en el contexto de {contexto}:\n{muestra}\n"
                  f"Genera UN título descriptivo de máximo 6 palabras que capture el tema central.\n"
                  f"Responde SOLO con el título, sin explicaciones.")
        try:
            resp = client.chat.complete(model="mistral-small-latest",
                                        messages=[{"role":"user","content":prompt}],
                                        max_tokens=20)
            etiquetas[grupo] = resp.choices[0].message.content.strip()
        except:
            etiquetas[grupo] = f"Grupo {grupo + 1}"
    return etiquetas

def categorizar_hallazgos(frases, contexto):
    texto = "\n".join(f"- {f}" for f in frases[:30])
    prompt = (f"Eres un experto en diagnóstico social y organizacional.\n"
              f"Analiza estas frases del contexto de {contexto}:\n{texto}\n\n"
              f"Clasifica los hallazgos en:\n"
              f"- PROBLEMAS: situaciones negativas\n"
              f"- POTENCIALIDADES: recursos o aspectos positivos\n"
              f"- PROPUESTAS: sugerencias planteadas por participantes\n"
              f"- ALERTAS: situaciones urgentes\n\n"
              f"Formato exacto:\nPROBLEMAS:\n- [hallazgo]\nPOTENCIALIDADES:\n- [hallazgo]\n"
              f"PROPUESTAS:\n- [hallazgo]\nALERTAS:\n- [hallazgo]\n\n"
              f"Máximo 3 puntos por categoría. Si no hay, escribe 'No identificados'.")
    try:
        resp = client.chat.complete(model="mistral-small-latest",
                                    messages=[{"role":"user","content":prompt}],
                                    max_tokens=500)
        return resp.choices[0].message.content.strip()
    except Exception as e:
        return f"Error: {e}"

def generar_informe_ia(resumen_datos, contexto):
    prompt = (f"Eres un consultor experto en {contexto}.\n"
              f"Basándote en este resumen de análisis semántico participativo:\n{resumen_datos}\n\n"
              f"Genera un informe ejecutivo profesional con:\n"
              f"1. DIAGNÓSTICO GENERAL (2-3 párrafos)\n"
              f"2. HALLAZGOS PRINCIPALES (máximo 5, por relevancia)\n"
              f"3. RECOMENDACIONES ESTRATÉGICAS (máximo 5, priorizadas)\n"
              f"4. ACCIONES INMEDIATAS (máximo 3 para los próximos 30 días)\n"
              f"5. ALERTAS (atención urgente)\n"
              f"6. NOTA SOBRE SUBREGISTRO\n\n"
              f"Lenguaje profesional, directo y orientado a toma de decisiones.")
    try:
        resp = client.chat.complete(model="mistral-large-latest",
                                    messages=[{"role":"user","content":prompt}],
                                    max_tokens=3500)
        return resp.choices[0].message.content.strip()
    except Exception as e:
        return f"Error: {e}"

def detectar_subregistro(df, cols_texto=None):
    alertas = []
    total = len(df)
    if total < 10:
        alertas.append(f"⚠️ Muestra pequeña: {total} registros.")
    if cols_texto:
        for col in cols_texto:
            if col in df.columns:
                vacias = df[col].isna().sum() + (df[col].astype(str) == "").sum()
                pct = round(vacias / total * 100, 1)
                if pct > 20:
                    alertas.append(f"⚠️ Columna '{col}': {pct}% vacías.")
        longs = df[cols_texto[0]].dropna().astype(str).str.len()
        if longs.mean() < 20:
            alertas.append("⚠️ Respuestas muy cortas en promedio.")
    if "cluster" in df.columns:
        dist = df["cluster"].value_counts(normalize=True)
        if dist.min() < 0.05:
            alertas.append("⚠️ Hay grupos con menos del 5% de respuestas.")
    if not alertas:
        alertas.append("✅ No se detectaron señales evidentes de subregistro.")
    return alertas

def normalizar_str(texto):
    """Convierte a minúsculas y elimina tildes/acentos para comparación robusta.
    Ej: 'Raíces del Futuro' → 'raices del futuro'"""
    texto = str(texto).strip().lower()
    return ''.join(
        c for c in unicodedata.normalize('NFD', texto)
        if unicodedata.category(c) != 'Mn'
    )


def normalizar_linea(texto):
    texto_lower = str(texto).lower().strip()
    for linea in LINEAS_CANONICAS:
        palabras_clave = linea.lower().split()
        if sum(1 for p in palabras_clave if p in texto_lower) >= len(palabras_clave) - 1:
            return linea
    for linea in LINEAS_CANONICAS:
        if any(p in texto_lower for p in linea.lower().split() if len(p) > 4):
            return linea
    return None


def texto_a_likert(valor, escala_max):
    """
    Convierte respuestas Likert de texto o número a float.
    Funciona con escalas 1-3 y 1-5. Retorna np.nan si no reconoce el valor.
    """
    if pd.isna(valor):
        return np.nan
    try:
        n = float(str(valor).strip().replace(",", "."))
        return n if 1 <= n <= escala_max else np.nan
    except (ValueError, TypeError):
        pass
    mapa3 = {
        "de acuerdo": 3, "sí": 3, "si": 3, "siempre": 3, "totalmente": 3,
        "muy de acuerdo": 3, "completamente de acuerdo": 3, "excelente": 3,
        "completamente": 3, "totalmente de acuerdo": 3,
        "neutral": 2, "a veces": 2, "parcialmente": 2, "regular": 2,
        "más o menos": 2, "más o menos de acuerdo": 2,
        "en desacuerdo": 1, "no": 1, "nunca": 1, "deficiente": 1,
        "muy en desacuerdo": 1, "totalmente en desacuerdo": 1,
        "completamente en desacuerdo": 1,
    }
    mapa5 = {
        "muy de acuerdo": 5, "totalmente de acuerdo": 5, "excelente": 5,
        "completamente de acuerdo": 5, "siempre": 5,
        "de acuerdo": 4, "bien": 4, "bastante": 4, "casi siempre": 4,
        "neutral": 3, "ni de acuerdo ni en desacuerdo": 3, "regular": 3,
        "a veces": 3,
        "en desacuerdo": 2, "poco": 2, "deficiente": 2, "casi nunca": 2,
        "muy en desacuerdo": 1, "totalmente en desacuerdo": 1,
        "pésimo": 1, "nunca": 1, "completamente en desacuerdo": 1,
    }
    mapa = mapa5 if escala_max == 5 else mapa3
    return mapa.get(str(valor).strip().lower(), np.nan)


def clasificar_texto_enc(texto):
    texto_lower = str(texto).lower()
    temas = []
    for tema, palabras in KEYWORDS_TEMAS.items():
        if any(p in texto_lower for p in palabras):
            temas.append(tema)
    return temas if temas else ["Sin clasificar"]

# ════════════════════════════════════════════════════════════════
# COLORES Y ESTILOS EXCEL (globales)
# ════════════════════════════════════════════════════════════════
C_AZUL1   = "1F3864"
C_AZUL2   = "2E75B6"
C_AZULBG  = "D6E4F0"
C_VERDE   = "1E7145"
C_VERDEBG = "E2EFDA"
C_AMARILLO= "BF8F00"
C_AMARILLBG="FFF2CC"
C_ROJO    = "C00000"
C_rojoBG  = "FDECEA"
C_GRIS    = "404040"

def xlsx_header(cell, bg=None, fg="FFFFFF", bold=True, sz=10, center=True):
    bg = bg or C_AZUL2
    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    cell.font = Font(color=fg, bold=bold, name="Calibri", size=sz)
    if center:
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def xlsx_data(cell, bg=None, bold=False, wrap=True, center=False):
    if bg:
        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    cell.font = Font(name="Calibri", size=10, bold=bold, color=C_GRIS)
    cell.alignment = Alignment(vertical="center", wrap_text=wrap,
                               horizontal="center" if center else "left")

def semaforo_cf(ws, rng, col_ref, verde_formula, amarillo_formula, rojo_formula):
    """Agrega formato condicional de semáforo a un rango."""
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[verde_formula],
        fill=PatternFill(start_color=C_VERDEBG, end_color=C_VERDEBG, fill_type="solid"),
        font=Font(color=C_VERDE, bold=True, name="Calibri")))
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[amarillo_formula],
        fill=PatternFill(start_color=C_AMARILLBG, end_color=C_AMARILLBG, fill_type="solid"),
        font=Font(color=C_AMARILLO, bold=True, name="Calibri")))
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[rojo_formula],
        fill=PatternFill(start_color=C_rojoBG, end_color=C_rojoBG, fill_type="solid"),
        font=Font(color=C_ROJO, bold=True, name="Calibri")))

# ════════════════════════════════════════════════════════════════
# BUILDER EXCEL — MÓDULO ENCUESTA CON FÓRMULAS
# ════════════════════════════════════════════════════════════════
def crear_excel_encuesta_formulas(dfs_ok, cfg, resultados):
    """
    Crea workbook openpyxl con fórmulas reales para el módulo Encuesta.
    cfg: dict con organizacion, cliente, municipios, periodo, escala_max,
         col_proyecto, cols_likert, cols_sino, col_cal, cols_texto, proyectos_lista
    resultados: dict con filas_consolidado, filas_tri, filas_cual, sin_clasificar
    """
    escala_max  = cfg.get("escala_max", 3)
    col_proy    = cfg.get("col_proyecto")
    cols_likert = cfg.get("cols_likert", [])
    cols_sino   = cfg.get("cols_sino", [])
    col_cal     = cfg.get("col_cal")
    cols_texto  = cfg.get("cols_texto", [])
    proyectos   = cfg.get("proyectos_lista", [])
    umbral_tri  = 0.5 if escala_max == 3 else 1.0

    # Filas de CONFIG (para fórmulas que referencian CONFIG)
    ROW_ESCALA  = 8   # B8 = valor escala_max
    ROW_VERDE   = 9   # B9 = =B8*0.9
    ROW_AMARI   = 10  # B10 = =B8*0.67

    wb = Workbook()
    wb.remove(wb.active)

    # ── 1. CONFIG ─────────────────────────────────────────────────
    ws_cfg = wb.create_sheet("CONFIG")
    ws_cfg.column_dimensions["A"].width = 32
    ws_cfg.column_dimensions["B"].width = 42

    secciones_cfg = [
        ("INFORMACIÓN GENERAL", None, True),
        ("Organización ejecutora", cfg.get("organizacion",""), False),
        ("Empresa cliente",        cfg.get("cliente",""),      False),
        ("Municipios / Territorios",cfg.get("municipios",""),  False),
        ("Período de evaluación",  cfg.get("periodo",""),      False),
        (None, None, False),
        ("UMBRALES DE SEMÁFORO (se actualizan automáticamente)", None, True),
        ("Escala máxima Likert",   escala_max,                 False),
        ("Umbral Verde (≥)",       "=B8*0.9",                  False),
        ("Umbral Amarillo (≥)",    "=B8*0.67",                 False),
        ("Umbral Rojo (<)",        "=B10",                     False),
        (None, None, False),
        ("PROYECTOS ANALIZADOS", None, True),
    ]

    for ri_c, (k, v, es_titulo) in enumerate(secciones_cfg, 1):
        if k:
            ws_cfg.cell(ri_c, 1).value = k
            if es_titulo:
                xlsx_header(ws_cfg.cell(ri_c, 1), bg=C_AZUL1, sz=11)
                ws_cfg.merge_cells(f"A{ri_c}:B{ri_c}")
            else:
                xlsx_data(ws_cfg.cell(ri_c, 1), bold=True)
        if v is not None:
            ws_cfg.cell(ri_c, 2).value = v
            xlsx_data(ws_cfg.cell(ri_c, 2))

    proy_start = len(secciones_cfg) + 1
    for j, proy in enumerate(proyectos):
        ws_cfg.cell(proy_start + j, 1).value = f"Proyecto {j+1}"
        ws_cfg.cell(proy_start + j, 2).value = proy
        xlsx_data(ws_cfg.cell(proy_start + j, 1), bold=True)
        xlsx_data(ws_cfg.cell(proy_start + j, 2))

    ws_cfg.freeze_panes = "B2"

    # ── 2. Hojas de datos crudos ──────────────────────────────────
    grupo_meta = {}   # {grupo: {col_name: col_idx, '_sheet': str, '_nrows': int}}

    for grupo, df_g in dfs_ok.items():
        # Nombre de hoja sin tildes ni caracteres especiales
        sname_safe = f"1_{grupo[:9]}".replace("á","a").replace("é","e")\
            .replace("í","i").replace("ó","o").replace("ú","u")\
            .replace("ñ","n").replace(" ","_")
        sname = sname_safe
        ws_g = wb.create_sheet(sname)
        cols_g = list(df_g.columns)
        col_map = {}

        for ci, col_h in enumerate(cols_g, 1):
            cell_h = ws_g.cell(1, ci)
            cell_h.value = col_h
            xlsx_header(cell_h)
            ws_g.column_dimensions[get_column_letter(ci)].width = max(14, len(str(col_h)) + 6)
            col_map[col_h] = ci

        for ri, (_, row_g) in enumerate(df_g.iterrows(), 2):
            for ci, col_h in enumerate(cols_g, 1):
                val = row_g[col_h]
                if pd.notna(val):
                    ws_g.cell(ri, ci).value = val
                xlsx_data(ws_g.cell(ri, ci))

        # Validación de datos para columnas Likert
        for cl in cols_likert:
            if cl in col_map:
                cltr = get_column_letter(col_map[cl])
                dv = DataValidation(
                    type="whole", operator="between",
                    formula1="1", formula2=str(escala_max),
                    showErrorMessage=True,
                    errorTitle="Valor inválido",
                    error=f"Ingresa un entero entre 1 y {escala_max}")
                ws_g.add_data_validation(dv)
                dv.add(f"{cltr}2:{cltr}1000")

        ws_g.freeze_panes = "A2"
        grupo_meta[grupo] = col_map
        grupo_meta[grupo]["_sheet"] = sname
        grupo_meta[grupo]["_nrows"] = len(df_g)

    # ── 3. 2_Likert: AVERAGEIFS + semáforo condicional ───────────
    if cols_likert and proyectos:
        ws_lk = wb.create_sheet("2_Likert")
        ws_lk.column_dimensions["A"].width = 36

        hdrs_lk = ["Proyecto"]
        for cl in cols_likert:
            for gr in dfs_ok:
                if cl in grupo_meta.get(gr, {}):
                    hdrs_lk.append(f"{cl} | {gr}")

        for ci, h in enumerate(hdrs_lk, 1):
            xlsx_header(ws_lk.cell(1, ci))
            ws_lk.cell(1, ci).value = h
            ws_lk.column_dimensions[get_column_letter(ci)].width = 22

        lk_col_letters = []
        for ri, proy in enumerate(proyectos, 2):
            ws_lk.cell(ri, 1).value = proy
            xlsx_data(ws_lk.cell(ri, 1), bold=True)
            ci_lk = 2
            for cl in cols_likert:
                for gr, gm in grupo_meta.items():
                    if cl not in gm or "_sheet" not in gm:
                        continue
                    sht     = gm["_sheet"]
                    cl_idx  = gm[cl]
                    cl_ltr  = get_column_letter(cl_idx)
                    nrows   = gm.get("_nrows", 1000)

                    if col_proy and col_proy in gm:
                        pry_ltr = get_column_letter(gm[col_proy])
                        f = (f"=IFERROR(AVERAGEIFS("
                             f"'{sht}'!{cl_ltr}$2:{cl_ltr}${nrows+1},"
                             f"'{sht}'!{pry_ltr}$2:{pry_ltr}${nrows+1},"
                             f"$A{ri}),\"\")")
                    else:
                        f = (f"=IFERROR(AVERAGE("
                             f"'{sht}'!{cl_ltr}$2:{cl_ltr}${nrows+1}),\"\")")

                    cell_lk = ws_lk.cell(ri, ci_lk)
                    cell_lk.value = f
                    cell_lk.number_format = "0.00"
                    xlsx_data(cell_lk)
                    lk_col_letters.append(get_column_letter(ci_lk))
                    ci_lk += 1

        n_proy = len(proyectos)
        for cl_cf in set(lk_col_letters):
            rng = f"{cl_cf}2:{cl_cf}{n_proy + 1}"
            semaforo_cf(ws_lk, rng, cl_cf,
                verde_formula  = f"{cl_cf}2>=CONFIG!$B${ROW_VERDE}",
                amarillo_formula = (f"AND({cl_cf}2>=CONFIG!$B${ROW_AMARI},"
                                    f"{cl_cf}2<CONFIG!$B${ROW_VERDE})"),
                rojo_formula   = (f"AND(ISNUMBER({cl_cf}2),"
                                  f"{cl_cf}2<CONFIG!$B${ROW_AMARI})"))

        nota_lk = n_proy + 3
        ws_lk.cell(nota_lk, 1).value = (
            "🟢 Destacado (≥ 90% escala)  |  🟡 Aceptable (67-89%)  |  "
            "🔴 Crítico (< 67%)  |  Umbrales se leen automáticamente desde CONFIG")
        ws_lk.cell(nota_lk, 1).font = Font(italic=True, color="777777", name="Calibri")
        ws_lk.freeze_panes = "B2"

    # ── 4. 3_SiNo_Cal: COUNTIFS + AVERAGEIFS calificación ────────
    if (cols_sino or col_cal) and proyectos:
        ws_sn = wb.create_sheet("3_SiNo_Cal")
        ws_sn.column_dimensions["A"].width = 36

        hdrs_sn = ["Proyecto"]
        for cs in cols_sino:
            for gr in dfs_ok:
                if cs in grupo_meta.get(gr, {}):
                    hdrs_sn.append(f"% Sí — {cs} | {gr}")
        if col_cal:
            for gr in dfs_ok:
                if col_cal in grupo_meta.get(gr, {}):
                    hdrs_sn.append(f"Calificación 1-5 | {gr}")

        for ci, h in enumerate(hdrs_sn, 1):
            xlsx_header(ws_sn.cell(1, ci))
            ws_sn.cell(1, ci).value = h
            ws_sn.column_dimensions[get_column_letter(ci)].width = 28

        cal_cols_cf = []
        for ri, proy in enumerate(proyectos, 2):
            ws_sn.cell(ri, 1).value = proy
            xlsx_data(ws_sn.cell(ri, 1), bold=True)
            ci_sn = 2

            for cs in cols_sino:
                for gr, gm in grupo_meta.items():
                    if cs not in gm or "_sheet" not in gm:
                        continue
                    sht    = gm["_sheet"]
                    cs_ltr = get_column_letter(gm[cs])
                    nrows  = gm.get("_nrows", 1000)

                    if col_proy and col_proy in gm:
                        py_ltr = get_column_letter(gm[col_proy])
                        # Cuenta "Sí" / total no vacío para ese proyecto
                        f_si = (f"=IFERROR("
                                f"(COUNTIFS('{sht}'!{py_ltr}$2:{py_ltr}${nrows+1},$A{ri},"
                                f"'{sht}'!{cs_ltr}$2:{cs_ltr}${nrows+1},\"Sí\")"
                                f"+COUNTIFS('{sht}'!{py_ltr}$2:{py_ltr}${nrows+1},$A{ri},"
                                f"'{sht}'!{cs_ltr}$2:{cs_ltr}${nrows+1},\"Si\"))"
                                f"/COUNTIFS('{sht}'!{py_ltr}$2:{py_ltr}${nrows+1},$A{ri},"
                                f"'{sht}'!{cs_ltr}$2:{cs_ltr}${nrows+1},\"<>\"),\"\")")
                    else:
                        f_si = (f"=IFERROR("
                                f"(COUNTIF('{sht}'!{cs_ltr}$2:{cs_ltr}${nrows+1},\"Sí\")"
                                f"+COUNTIF('{sht}'!{cs_ltr}$2:{cs_ltr}${nrows+1},\"Si\"))"
                                f"/COUNTA('{sht}'!{cs_ltr}$2:{cs_ltr}${nrows+1}),\"\")")

                    cell_sn = ws_sn.cell(ri, ci_sn)
                    cell_sn.value = f_si
                    cell_sn.number_format = "0%"
                    xlsx_data(cell_sn)
                    ci_sn += 1

            if col_cal:
                for gr, gm in grupo_meta.items():
                    if col_cal not in gm or "_sheet" not in gm:
                        continue
                    sht     = gm["_sheet"]
                    cal_ltr = get_column_letter(gm[col_cal])
                    nrows   = gm.get("_nrows", 1000)

                    if col_proy and col_proy in gm:
                        py_ltr = get_column_letter(gm[col_proy])
                        f_cal = (f"=IFERROR(AVERAGEIFS("
                                 f"'{sht}'!{cal_ltr}$2:{cal_ltr}${nrows+1},"
                                 f"'{sht}'!{py_ltr}$2:{py_ltr}${nrows+1},$A{ri}),\"\")")
                    else:
                        f_cal = (f"=IFERROR(AVERAGE("
                                 f"'{sht}'!{cal_ltr}$2:{cal_ltr}${nrows+1}),\"\")")

                    cell_cal = ws_sn.cell(ri, ci_sn)
                    cell_cal.value = f_cal
                    cell_cal.number_format = "0.00"
                    xlsx_data(cell_cal)
                    cal_cols_cf.append((get_column_letter(ci_sn), len(proyectos)))
                    ci_sn += 1

        for cl_cal_cf, n_p in cal_cols_cf:
            rng_cal = f"{cl_cal_cf}2:{cl_cal_cf}{n_p + 1}"
            semaforo_cf(ws_sn, rng_cal, cl_cal_cf,
                verde_formula   = f"{cl_cal_cf}2>=4.5",
                amarillo_formula = f"AND({cl_cal_cf}2>=3.5,{cl_cal_cf}2<4.5)",
                rojo_formula    = f"AND(ISNUMBER({cl_cal_cf}2),{cl_cal_cf}2<3.5)")

        ws_sn.freeze_panes = "B2"

    # ── 5. 4_Triangulacion: MAX-MIN + Convergencia ────────────────
    if cols_likert and len(dfs_ok) >= 2 and proyectos:
        ws_tri = wb.create_sheet("4_Triangulacion")
        grupos_t = list(dfs_ok.keys())
        hdrs_t = ["Proyecto", "Indicador"] + grupos_t + ["Rango MAX-MIN", "Convergencia"]
        for ci, h in enumerate(hdrs_t, 1):
            xlsx_header(ws_tri.cell(1, ci))
            ws_tri.cell(1, ci).value = h
            ws_tri.column_dimensions[get_column_letter(ci)].width = (
                36 if ci == 1 else (28 if ci == 2 else 18))

        ri_t = 2
        for proy in proyectos:
            for cl in cols_likert:
                ws_tri.cell(ri_t, 1).value = proy
                ws_tri.cell(ri_t, 2).value = cl
                xlsx_data(ws_tri.cell(ri_t, 1), bold=True)
                xlsx_data(ws_tri.cell(ri_t, 2))

                ci_t = 3
                grupo_val_refs = []
                for gr in grupos_t:
                    gm = grupo_meta.get(gr, {})
                    if cl in gm and "_sheet" in gm:
                        sht    = gm["_sheet"]
                        cl_ltr = get_column_letter(gm[cl])
                        nrows  = gm.get("_nrows", 1000)
                        if col_proy and col_proy in gm:
                            py_ltr = get_column_letter(gm[col_proy])
                            f_t = (f"=IFERROR(AVERAGEIFS("
                                   f"'{sht}'!{cl_ltr}$2:{cl_ltr}${nrows+1},"
                                   f"'{sht}'!{py_ltr}$2:{py_ltr}${nrows+1},"
                                   f"$A{ri_t}),\"\")")
                        else:
                            f_t = (f"=IFERROR(AVERAGE("
                                   f"'{sht}'!{cl_ltr}$2:{cl_ltr}${nrows+1}),\"\")")
                        cell_t = ws_tri.cell(ri_t, ci_t)
                        cell_t.value = f_t
                        cell_t.number_format = "0.00"
                        xlsx_data(cell_t)
                        grupo_val_refs.append(f"{get_column_letter(ci_t)}{ri_t}")
                    else:
                        ws_tri.cell(ri_t, ci_t).value = "Sin dato"
                        xlsx_data(ws_tri.cell(ri_t, ci_t), bg=C_AZULBG)
                    ci_t += 1

                # Rango MAX-MIN
                if grupo_val_refs:
                    refs = ",".join(grupo_val_refs)
                    rango_ci = ci_t
                    rango_ltr = get_column_letter(rango_ci)
                    cell_rng = ws_tri.cell(ri_t, rango_ci)
                    cell_rng.value = f"=IFERROR(MAX({refs})-MIN({refs}),\"\")"
                    cell_rng.number_format = "0.00"
                    xlsx_data(cell_rng)
                    ci_t += 1

                    # Convergencia
                    cell_conv = ws_tri.cell(ri_t, ci_t)
                    cell_conv.value = (f"=IFERROR(IF({rango_ltr}{ri_t}<={umbral_tri},"
                                       f'"✅ Convergencia","⚠️ Divergencia"),"")')
                    xlsx_data(cell_conv)

                    conv_ltr = get_column_letter(ci_t)
                    n_p_t = len(proyectos) * len(cols_likert)
                    rng_cv = f"{conv_ltr}2:{conv_ltr}{n_p_t + 1}"
                    ws_tri.conditional_formatting.add(rng_cv, FormulaRule(
                        formula=[f'{conv_ltr}{ri_t}="✅ Convergencia"'],
                        fill=PatternFill(start_color=C_VERDEBG, end_color=C_VERDEBG,
                                        fill_type="solid")))
                    ws_tri.conditional_formatting.add(rng_cv, FormulaRule(
                        formula=[f'{conv_ltr}{ri_t}="⚠️ Divergencia"'],
                        fill=PatternFill(start_color=C_rojoBG, end_color=C_rojoBG,
                                        fill_type="solid")))

                ri_t += 1

        nota_t = ri_t + 1
        ws_tri.cell(nota_t, 1).value = (
            f"Umbral de convergencia: diferencia MAX-MIN ≤ {umbral_tri} "
            f"({'escala 1-3' if escala_max == 3 else 'escala 1-5'})")
        ws_tri.cell(nota_t, 1).font = Font(italic=True, color="777777", name="Calibri")
        ws_tri.freeze_panes = "C2"

    # ── 6. 5_Abiertas: datos cualitativos + desplegable ──────────
    if cols_texto:
        ws_ab = wb.create_sheet("5_Abiertas")
        ws_ab.column_dimensions["A"].width = 16
        ws_ab.column_dimensions["B"].width = 22
        ws_ab.column_dimensions["C"].width = 22
        ws_ab.column_dimensions["D"].width = 60
        ws_ab.column_dimensions["E"].width = 32

        hdrs_ab = ["Grupo", "Proyecto", "Columna", "Respuesta", "Tema (editar)"]
        for ci, h in enumerate(hdrs_ab, 1):
            xlsx_header(ws_ab.cell(1, ci))
            ws_ab.cell(1, ci).value = h

        # DataValidation: máx ~255 chars en la fórmula de lista
        # Usar solo los primeros temas para no corromper el Excel
        temas_lista = list(KEYWORDS_TEMAS.keys())[:8] + ["Sin clasificar"]
        temas_str = ",".join(temas_lista)
        # Verificar longitud (Excel tiene límite de ~255 caracteres)
        if len(temas_str) > 250:
            temas_lista = ["Agua y medio ambiente","Liderazgo y organización",
                           "Educación y formación","Infraestructura",
                           "Continuidad y sostenibilidad","Mejora y sugerencias",
                           "Percepción positiva empresa","Sin clasificar"]
            temas_str = ",".join(temas_lista)
        dv_temas = DataValidation(type="list", formula1=f'"{temas_str}"',
                                  showErrorMessage=False, sqref="E2:E5000")
        ws_ab.add_data_validation(dv_temas)

        ri_ab = 2
        for gr, df_g in dfs_ok.items():
            for ct in cols_texto:
                if ct not in df_g.columns:
                    continue
                for _, row_ab in df_g.iterrows():
                    txt = str(row_ab.get(ct, ""))
                    if txt and txt != "nan" and len(txt.strip()) > 2:
                        proy_ab = ""
                        if col_proy and col_proy in df_g.columns:
                            proy_ab = str(row_ab.get(col_proy, ""))
                        # Usar normalizar_str para eliminar tildes — igual que la app
                        tema_sug = "Sin clasificar"
                        txt_norm = normalizar_str(txt)
                        for tema_k, pals in KEYWORDS_TEMAS.items():
                            if any(normalizar_str(p) in txt_norm for p in pals):
                                tema_sug = tema_k
                                break
                        ws_ab.cell(ri_ab, 1).value = gr
                        ws_ab.cell(ri_ab, 2).value = proy_ab
                        ws_ab.cell(ri_ab, 3).value = ct
                        ws_ab.cell(ri_ab, 4).value = txt
                        ws_ab.cell(ri_ab, 5).value = tema_sug
                        ws_ab.cell(ri_ab, 4).alignment = Alignment(wrap_text=True,
                                                                    vertical="top")
                        # (validación asignada por rango en sqref, no por celda)
                        # Color Sin clasificar
                        if tema_sug == "Sin clasificar":
                            ws_ab.cell(ri_ab, 5).fill = PatternFill(
                                start_color=C_AMARILLBG, end_color=C_AMARILLBG,
                                fill_type="solid")
                        ri_ab += 1

        ws_ab.freeze_panes = "A2"

    # ── 7. 6_Dashboard ───────────────────────────────────────────
    ws_dash = wb.create_sheet("6_Dashboard")
    ws_dash.column_dimensions["A"].width = 34
    ws_dash.column_dimensions["B"].width = 24
    ws_dash.column_dimensions["C"].width = 24
    ws_dash.column_dimensions["D"].width = 24

    # Título
    tc = ws_dash.cell(1, 1)
    tc.value = "DASHBOARD EJECUTIVO — ENCUESTA DE INVERSIÓN SOCIAL"
    tc.font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    tc.fill = PatternFill(start_color=C_AZUL1, end_color=C_AZUL1, fill_type="solid")
    tc.alignment = Alignment(horizontal="center")
    ws_dash.merge_cells("A1:D1")
    ws_dash.row_dimensions[1].height = 28

    meta_fields = [
        ("Organización:", "=CONFIG!B2"),
        ("Empresa cliente:", "=CONFIG!B3"),
        ("Municipios:", "=CONFIG!B4"),
        ("Período:", "=CONFIG!B5"),
        ("Escala Likert (máximo):", "=CONFIG!B8"),
        ("Umbral Verde (≥):", f"=CONFIG!B{ROW_VERDE}"),
        ("Umbral Amarillo (≥):", f"=CONFIG!B{ROW_AMARI}"),
    ]
    for ri_d, (lbl, val) in enumerate(meta_fields, 3):
        ws_dash.cell(ri_d, 1).value = lbl
        ws_dash.cell(ri_d, 2).value = val
        xlsx_data(ws_dash.cell(ri_d, 1), bold=True)
        xlsx_data(ws_dash.cell(ri_d, 2))

    ri_d_hdr = 11
    xlsx_header(ws_dash.cell(ri_d_hdr, 1), bg=C_AZUL2)
    ws_dash.cell(ri_d_hdr, 1).value = "RESUMEN POR PROYECTO"
    ws_dash.merge_cells(f"A{ri_d_hdr}:D{ri_d_hdr}")

    hdrs_dash = ["Proyecto", "Grupos", "Indicadores Likert", "Detalle en hoja"]
    for ci, h in enumerate(hdrs_dash, 1):
        xlsx_header(ws_dash.cell(ri_d_hdr + 1, ci), bg=C_AZUL2)
        ws_dash.cell(ri_d_hdr + 1, ci).value = h

    for j, proy in enumerate(proyectos, ri_d_hdr + 2):
        ws_dash.cell(j, 1).value = proy
        ws_dash.cell(j, 2).value = ", ".join(dfs_ok.keys())
        ws_dash.cell(j, 3).value = ", ".join(cols_likert) if cols_likert else "—"
        ws_dash.cell(j, 4).value = "Ver 2_Likert · 3_SiNo_Cal · 4_Triangulacion"
        for ci_d in range(1, 5):
            xlsx_data(ws_dash.cell(j, ci_d))

    # ── 8. 7_Guia ────────────────────────────────────────────────
    ws_guia = wb.create_sheet("7_Guia")
    ws_guia.column_dimensions["A"].width = 85

    xlsx_header(ws_guia.cell(1, 1), bg=C_AZUL1, sz=13)
    ws_guia.cell(1, 1).value = "GUÍA DE USO — SISTEMA EXCEL ENCUESTA"

    guia = [
        "",
        "HOJAS DE ESTE ARCHIVO",
        "  CONFIG          → Cerebro del sistema. Modifica aquí los umbrales; todo se actualiza solo.",
        "  1_Comunidad / 1_Aliados / 1_Empresa → Datos crudos con validación automática de Likert.",
        "  2_Likert         → Promedios AVERAGEIFS por proyecto. Semáforo 🟢🟡🔴 automático.",
        "  3_SiNo_Cal       → COUNTIFS para Sí/No. AVERAGEIFS para calificación 1-5.",
        "  4_Triangulacion  → MAX-MIN entre grupos. ✅ Convergencia / ⚠️ Divergencia automática.",
        "  5_Abiertas       → Respuestas cualitativas. Usa el desplegable en columna E para reclasificar.",
        "  6_Dashboard      → Resumen ejecutivo con fórmulas que leen desde CONFIG.",
        "",
        "CÓMO FUNCIONAN LOS SEMÁFOROS",
        f"  Escala actual: 1 a {escala_max}",
        f"  🟢 Destacado  = promedio ≥ {escala_max * 0.9:.2f}  (90% de la escala)",
        f"  🟡 Aceptable  = promedio {escala_max * 0.67:.2f}–{escala_max * 0.9 - 0.01:.2f}  (67-89%)",
        f"  🔴 Crítico    = promedio < {escala_max * 0.67:.2f}  (menos del 67%)",
        "  Si cambias la escala en CONFIG!B8, los umbrales y colores se recalculan solos.",
        "",
        "CÓMO AGREGAR DATOS NUEVOS",
        "  1. Agrega filas al final de las hojas 1_Comunidad / 1_Aliados / 1_Empresa.",
        "  2. Los promedios en 2_Likert, 3_SiNo_Cal y 4_Triangulacion se actualizan automáticamente.",
        "  3. NO modifiques las fórmulas de las hojas 2, 3 y 4 — solo agrega datos en las hojas 1.",
        "",
        "TRIANGULACIÓN",
        f"  Umbral de convergencia: diferencia MAX-MIN ≤ {umbral_tri}",
        "  Si solo un grupo tiene datos, la triangulación muestra 'Sin dato' en los grupos faltantes.",
        "",
        "RESPUESTAS ABIERTAS (hoja 5_Abiertas)",
        "  Columna E = tema sugerido automáticamente. Puedes cambiar el tema con el desplegable.",
        "  Las respuestas amarillas son 'Sin clasificar' — revísalas manualmente.",
        "",
        "Generado automáticamente por el Módulo de Encuesta — Análisis Semántico.",
    ]
    for i, line in enumerate(guia, 2):
        ws_guia.cell(i, 1).value = line
        if line and not line.startswith(" ") and line.isupper():
            ws_guia.cell(i, 1).font = Font(bold=True, color=C_AZUL1, name="Calibri")
        elif line.startswith("  ") and line.strip().startswith("🟢"):
            ws_guia.cell(i, 1).fill = PatternFill(start_color=C_VERDEBG,
                                                   end_color=C_VERDEBG, fill_type="solid")
        elif line.startswith("  ") and line.strip().startswith("🟡"):
            ws_guia.cell(i, 1).fill = PatternFill(start_color=C_AMARILLBG,
                                                   end_color=C_AMARILLBG, fill_type="solid")
        elif line.startswith("  ") and line.strip().startswith("🔴"):
            ws_guia.cell(i, 1).fill = PatternFill(start_color=C_rojoBG,
                                                   end_color=C_rojoBG, fill_type="solid")
        else:
            ws_guia.cell(i, 1).font = Font(name="Calibri", size=10)

    return wb


# ════════════════════════════════════════════════════════════════
# BUILDER EXCEL — MÓDULO CARTOGRAFÍA CON FÓRMULAS
# ════════════════════════════════════════════════════════════════
def crear_excel_cartografia_formulas(df_filtrado, comp_filtro, cat_por_componente,
                                     cat_por_vereda, conteo_comp, pesos_c, resumen_exec,
                                     contexto="diagnóstico territorial participativo"):
    """Crea workbook openpyxl con fórmulas y formato para Cartografía Social."""
    wb = Workbook()
    wb.remove(wb.active)

    # ── CONFIG Cartografía ────────────────────────────────────────
    ws_cfg = wb.create_sheet("CONFIG")
    ws_cfg.column_dimensions["A"].width = 32
    ws_cfg.column_dimensions["B"].width = 42
    xlsx_header(ws_cfg.cell(1, 1), bg=C_AZUL1, sz=12)
    ws_cfg.cell(1, 1).value = "CONFIGURACIÓN — CARTOGRAFÍA SOCIAL"
    ws_cfg.merge_cells("A1:B1")

    cfg_cart = [
        ("Contexto del análisis", contexto),
        ("Componentes analizados", ", ".join(comp_filtro)),
        ("Total frases", len(df_filtrado)),
        ("Umbral similitud semántica", 0.75),
        ("Máx. frases representativas", 5),
        ("Líneas canónicas",
         " | ".join(LINEAS_CANONICAS)),
    ]
    for ri_c, (k, v) in enumerate(cfg_cart, 2):
        ws_cfg.cell(ri_c, 1).value = k
        ws_cfg.cell(ri_c, 2).value = v
        xlsx_data(ws_cfg.cell(ri_c, 1), bold=True)
        xlsx_data(ws_cfg.cell(ri_c, 2))

    # ── Datos completos ───────────────────────────────────────────
    ws_dat = wb.create_sheet("Datos completos")
    cols_meta_e = [c for c in ["municipio","vereda","año","semestre"]
                   if c in df_filtrado.columns]
    cols_dat = cols_meta_e + ["componente","frase","grupo","peso_semantico","lineas_inversion"]
    cols_dat = [c for c in cols_dat if c in df_filtrado.columns]
    for ci, h in enumerate(cols_dat, 1):
        xlsx_header(ws_dat.cell(1, ci))
        ws_dat.cell(1, ci).value = h
        ws_dat.column_dimensions[get_column_letter(ci)].width = max(14, len(h) + 6)
    for ri_d, (_, row_d) in enumerate(df_filtrado[cols_dat].iterrows(), 2):
        for ci, col_d in enumerate(cols_dat, 1):
            val_d = row_d[col_d]
            if pd.notna(val_d):
                ws_dat.cell(ri_d, ci).value = val_d
            xlsx_data(ws_dat.cell(ri_d, ci))
            # Formato condicional en peso_semantico
    if "peso_semantico" in cols_dat:
        ps_ci = cols_dat.index("peso_semantico") + 1
        ps_ltr = get_column_letter(ps_ci)
        n_dat = len(df_filtrado)
        rng_ps = f"{ps_ltr}2:{ps_ltr}{n_dat + 1}"
        semaforo_cf(ws_dat, rng_ps, ps_ltr,
            verde_formula   = f"{ps_ltr}2>=0.85",
            amarillo_formula = f"AND({ps_ltr}2>=0.65,{ps_ltr}2<0.85)",
            rojo_formula    = f"AND(ISNUMBER({ps_ltr}2),{ps_ltr}2<0.65)")
    ws_dat.freeze_panes = "A2"

    # ── Distribución componentes ──────────────────────────────────
    ws_dc = wb.create_sheet("Distribucion componentes")
    for ci, h in enumerate(conteo_comp.columns, 1):
        xlsx_header(ws_dc.cell(1, ci))
        ws_dc.cell(1, ci).value = h
        ws_dc.column_dimensions[get_column_letter(ci)].width = 28
    for ri_c, (_, row_c) in enumerate(conteo_comp.iterrows(), 2):
        for ci, col_c in enumerate(conteo_comp.columns, 1):
            ws_dc.cell(ri_c, ci).value = row_c[col_c]
            xlsx_data(ws_dc.cell(ri_c, ci))
    # Fórmulas: total y porcentaje
    n_comp = len(conteo_comp)
    ws_dc.cell(n_comp + 2, 1).value = "TOTAL"
    ws_dc.cell(n_comp + 2, 1).font = Font(bold=True, name="Calibri")
    ws_dc.cell(n_comp + 2, 2).value = f"=SUM(B2:B{n_comp + 1})"
    ws_dc.cell(n_comp + 2, 2).font = Font(bold=True, name="Calibri")
    # Columna porcentaje
    xlsx_header(ws_dc.cell(1, 3), bg=C_AZULBG, fg=C_GRIS)
    ws_dc.cell(1, 3).value = "Porcentaje"
    ws_dc.column_dimensions["C"].width = 16
    for ri_pct in range(2, n_comp + 2):
        ws_dc.cell(ri_pct, 3).value = f"=B{ri_pct}/B{n_comp + 2}"
        ws_dc.cell(ri_pct, 3).number_format = "0.0%"

    # ── Líneas de inversión ───────────────────────────────────────
    lineas_exp = []
    for ls in df_filtrado["lineas_inversion"]:
        for l in str(ls).split(","):
            norm = normalizar_linea(l.strip().rstrip(".").strip())
            if norm:
                lineas_exp.append(norm)
    if lineas_exp:
        ws_li = wb.create_sheet("Lineas de inversion")
        df_li = pd.DataFrame({"linea": lineas_exp})
        cnt_li = df_li["linea"].value_counts().reset_index()
        cnt_li.columns = ["Línea de inversión", "Frecuencia"]
        for ci, h in enumerate(cnt_li.columns, 1):
            xlsx_header(ws_li.cell(1, ci))
            ws_li.cell(1, ci).value = h
            ws_li.column_dimensions[get_column_letter(ci)].width = 32
        for ri_li, (_, row_li) in enumerate(cnt_li.iterrows(), 2):
            for ci, col_li in enumerate(cnt_li.columns, 1):
                ws_li.cell(ri_li, ci).value = row_li[col_li]
                xlsx_data(ws_li.cell(ri_li, ci))
        n_li = len(cnt_li)
        ws_li.cell(n_li + 2, 1).value = "TOTAL"
        ws_li.cell(n_li + 2, 1).font = Font(bold=True, name="Calibri")
        ws_li.cell(n_li + 2, 2).value = f"=SUM(B2:B{n_li + 1})"
        ws_li.cell(n_li + 2, 2).font = Font(bold=True, name="Calibri")

    # ── Cruce componente x línea ──────────────────────────────────
    cruce_rows = []
    for _, row_cr in df_filtrado.iterrows():
        for l in str(row_cr["lineas_inversion"]).split(","):
            norm = normalizar_linea(l.strip().rstrip(".").strip())
            if norm:
                cruce_rows.append({"Componente": row_cr["componente"], "Línea": norm})
    if cruce_rows:
        ws_cx = wb.create_sheet("Cruce componente x linea")
        df_cx = pd.DataFrame(cruce_rows)
        piv_cx = df_cx.groupby(["Componente","Línea"]).size().reset_index(name="Cantidad")
        for ci, h in enumerate(piv_cx.columns, 1):
            xlsx_header(ws_cx.cell(1, ci))
            ws_cx.cell(1, ci).value = h
            ws_cx.column_dimensions[get_column_letter(ci)].width = 30
        for ri_cx, (_, row_cx) in enumerate(piv_cx.iterrows(), 2):
            for ci, col_cx in enumerate(piv_cx.columns, 1):
                ws_cx.cell(ri_cx, ci).value = row_cx[col_cx]
                xlsx_data(ws_cx.cell(ri_cx, ci))

    # ── Grupos por componente ─────────────────────────────────────
    ws_gp = wb.create_sheet("Grupos por componente")
    gp_hdrs = ["Componente","Grupo","Frecuencia","Frase representativa"]
    for ci, h in enumerate(gp_hdrs, 1):
        xlsx_header(ws_gp.cell(1, ci))
        ws_gp.cell(1, ci).value = h
        ws_gp.column_dimensions[get_column_letter(ci)].width = (
            24 if ci < 3 else (16 if ci == 3 else 60))
    ri_gp = 2
    for comp_gp in comp_filtro:
        sub_gp = df_filtrado[df_filtrado["componente"] == comp_gp]
        for grp_gp in sub_gp["grupo"].unique():
            sub_g = sub_gp[sub_gp["grupo"] == grp_gp]
            frase_r = sub_g.nlargest(1, "peso_semantico").iloc[0]["frase"]
            ws_gp.cell(ri_gp, 1).value = comp_gp
            ws_gp.cell(ri_gp, 2).value = grp_gp
            ws_gp.cell(ri_gp, 3).value = len(sub_g)
            ws_gp.cell(ri_gp, 4).value = frase_r
            ws_gp.cell(ri_gp, 4).alignment = Alignment(wrap_text=True, vertical="top")
            for ci in range(1, 5):
                xlsx_data(ws_gp.cell(ri_gp, ci))
            ri_gp += 1

    # ── Cohesión semántica con semáforo ──────────────────────────
    ws_coh = wb.create_sheet("Cohesion semantica")
    for ci, h in enumerate(pesos_c.columns, 1):
        xlsx_header(ws_coh.cell(1, ci))
        ws_coh.cell(1, ci).value = h
        ws_coh.column_dimensions[get_column_letter(ci)].width = 28
    for ri_coh, (_, row_coh) in enumerate(pesos_c.iterrows(), 2):
        for ci, col_coh in enumerate(pesos_c.columns, 1):
            ws_coh.cell(ri_coh, ci).value = row_coh[col_coh]
            xlsx_data(ws_coh.cell(ri_coh, ci))
    # Semáforo en columna de cohesión (segunda columna)
    n_coh = len(pesos_c)
    semaforo_cf(ws_coh, f"B2:B{n_coh + 1}", "B",
        verde_formula   = "B2>=0.75",
        amarillo_formula = "AND(B2>=0.55,B2<0.75)",
        rojo_formula    = "AND(ISNUMBER(B2),B2<0.55)")
    nota_coh = n_coh + 2
    ws_coh.cell(nota_coh, 1).value = (
        "🟢 Alta cohesión (≥0.75) | 🟡 Media (0.55-0.74) | 🔴 Baja (<0.55)")
    ws_coh.cell(nota_coh, 1).font = Font(italic=True, color="777777", name="Calibri")

    # ── Resumen ejecutivo ─────────────────────────────────────────
    ws_re = wb.create_sheet("Resumen ejecutivo")
    for ci, h in enumerate(resumen_exec.columns, 1):
        xlsx_header(ws_re.cell(1, ci))
        ws_re.cell(1, ci).value = h
        ws_re.column_dimensions[get_column_letter(ci)].width = 24
    for ri_re, (_, row_re) in enumerate(resumen_exec.iterrows(), 2):
        for ci, col_re in enumerate(resumen_exec.columns, 1):
            ws_re.cell(ri_re, ci).value = row_re[col_re]
            xlsx_data(ws_re.cell(ri_re, ci))
    n_re = len(resumen_exec)
    ws_re.cell(n_re + 2, 1).value = "TOTAL FRASES"
    ws_re.cell(n_re + 2, 1).font = Font(bold=True, name="Calibri")
    ws_re.cell(n_re + 2, 2).value = f"=SUM(B2:B{n_re + 1})"
    ws_re.cell(n_re + 2, 2).font = Font(bold=True, name="Calibri")

    # ── Frases representativas ────────────────────────────────────
    ws_fr = wb.create_sheet("Frases representativas")
    fr_hdrs = ["Componente","Frase","Peso","Relevancia","Frases similares",
               "Variaciones","Líneas de inversión"]
    for ci, h in enumerate(fr_hdrs, 1):
        xlsx_header(ws_fr.cell(1, ci))
        ws_fr.cell(1, ci).value = h
        ws_fr.column_dimensions[get_column_letter(ci)].width = (
            22 if ci in [1,3,4,5] else 55)

    ri_fr = 2
    for comp_fr in comp_filtro:
        sub_fr = df_filtrado[df_filtrado["componente"] == comp_fr]
        top_fr = sub_fr.nlargest(5, "peso_semantico")
        frases_v, vecs_v = [], []
        for _, row_fr in top_fr.iterrows():
            frase_fr = row_fr["frase"].strip()
            if vecs_v:
                v = modelo.encode([frase_fr])
                sims_fr = cosine_similarity(v, vecs_v)[0]
                if any(s > 0.75 for s in sims_fr):
                    continue
            peso_fr = round(float(row_fr["peso_semantico"]), 3)
            rel_fr = "Alta 🔴" if peso_fr >= 0.85 else ("Media 🟡" if peso_fr >= 0.65 else "Baja 🟢")
            grp_fr = row_fr["grupo"]
            sub_g_fr = sub_fr[sub_fr["grupo"] == grp_fr]
            otras_fr = sub_g_fr[sub_g_fr["frase"] != frase_fr]["frase"].head(2).tolist()
            ws_fr.cell(ri_fr, 1).value = comp_fr
            ws_fr.cell(ri_fr, 2).value = frase_fr
            ws_fr.cell(ri_fr, 3).value = peso_fr
            ws_fr.cell(ri_fr, 4).value = rel_fr
            ws_fr.cell(ri_fr, 5).value = sub_g_fr.shape[0]
            ws_fr.cell(ri_fr, 6).value = " | ".join(otras_fr)
            ws_fr.cell(ri_fr, 7).value = str(row_fr["lineas_inversion"])
            for ci in range(1, 8):
                xlsx_data(ws_fr.cell(ri_fr, ci))
                ws_fr.cell(ri_fr, ci).alignment = Alignment(wrap_text=True, vertical="top")
            frases_v.append(frase_fr)
            vecs_v.append(modelo.encode([frase_fr])[0])
            ri_fr += 1

    # ── Tabla detallada ───────────────────────────────────────────
    ws_td = wb.create_sheet("Tabla detallada")
    cols_td = [c for c in ["municipio","vereda","año","semestre","componente",
                            "frase","grupo","peso_semantico","lineas_inversion"]
               if c in df_filtrado.columns]
    for ci, h in enumerate(cols_td, 1):
        xlsx_header(ws_td.cell(1, ci))
        ws_td.cell(1, ci).value = h
        ws_td.column_dimensions[get_column_letter(ci)].width = max(14, len(h) + 6)
    for ri_td, (_, row_td) in enumerate(df_filtrado[cols_td].iterrows(), 2):
        for ci, col_td in enumerate(cols_td, 1):
            val_td = row_td[col_td]
            if pd.notna(val_td):
                ws_td.cell(ri_td, ci).value = val_td
            xlsx_data(ws_td.cell(ri_td, ci))
    ws_td.freeze_panes = "A2"

    # ── Categorización por componente ─────────────────────────────
    ws_ccat = wb.create_sheet("Categorizacion componente")
    for ci, h in enumerate(["Vereda","Componente","Categorización"], 1):
        xlsx_header(ws_ccat.cell(1, ci))
        ws_ccat.cell(1, ci).value = h
        ws_ccat.column_dimensions[get_column_letter(ci)].width = (16 if ci<3 else 80)
    ri_ccat = 2
    for comp_cc, txt_cc in cat_por_componente.items():
        for linea_cc in txt_cc.split("\n"):
            if linea_cc.strip():
                ws_ccat.cell(ri_ccat, 1).value = "TODAS"
                ws_ccat.cell(ri_ccat, 2).value = comp_cc
                ws_ccat.cell(ri_ccat, 3).value = linea_cc.strip()
                ws_ccat.cell(ri_ccat, 3).alignment = Alignment(wrap_text=True,vertical="top")
                ri_ccat += 1

    # ── Categorización por vereda ─────────────────────────────────
    ws_cver = wb.create_sheet("Categorizacion por vereda")
    for ci, h in enumerate(["Vereda","Componente","Categorización"], 1):
        xlsx_header(ws_cver.cell(1, ci))
        ws_cver.cell(1, ci).value = h
        ws_cver.column_dimensions[get_column_letter(ci)].width = (20 if ci<3 else 80)
    ri_cver = 2
    for comp_cv, veredas_cv in cat_por_vereda.items():
        for vereda_cv, txt_cv in veredas_cv.items():
            for linea_cv in txt_cv.split("\n"):
                if linea_cv.strip():
                    ws_cver.cell(ri_cver, 1).value = vereda_cv
                    ws_cver.cell(ri_cver, 2).value = comp_cv
                    ws_cver.cell(ri_cver, 3).value = linea_cv.strip()
                    ws_cver.cell(ri_cver, 3).alignment = Alignment(wrap_text=True,vertical="top")
                    ri_cver += 1

    return wb


# ════════════════════════════════════════════════════════════════
# SELECTOR DE MODO
# ════════════════════════════════════════════════════════════════
modo = st.sidebar.radio("Selecciona el tipo de análisis:",
                        ["📋 Encuesta", "🗺️ Cartografía Social"])

# ════════════════════════════════════════════════════════════════
# MÓDULO 1 — ENCUESTA
# ════════════════════════════════════════════════════════════════
if modo == "📋 Encuesta":
    st.header("📋 Módulo de Encuesta")
    st.markdown("Análisis de encuestas de inversión social — tres grupos: "
                "**Comunidad**, **Aliados** y **Empresa**.")

    def semaforo_likert(valor, escala_max):
        if pd.isna(valor):
            return "⬜ Sin dato"
        if valor >= escala_max * 0.9:
            return f"🟢 {valor:.2f}"
        elif valor >= escala_max * 0.67:
            return f"🟡 {valor:.2f}"
        return f"🔴 {valor:.2f}"

    def semaforo_rating(valor):
        if pd.isna(valor):
            return "⬜ Sin dato"
        if valor >= 4.5:
            return f"🟢 {valor:.2f}"
        elif valor >= 3.5:
            return f"🟡 {valor:.2f}"
        return f"🔴 {valor:.2f}"

    def convergencia(rango, escala_max):
        return "✅ Convergencia" if rango <= (0.5 if escala_max == 3 else 1.0) else "⚠️ Divergencia"

    # ── Configuración general ─────────────────────────────────────
    with st.expander("⚙️ Configuración del análisis", expanded=True):
        c1, c2 = st.columns(2)
        with c1:
            org_enc    = st.text_input("Organización ejecutora:", placeholder="Ej: CDC")
            cli_enc    = st.text_input("Empresa cliente:", placeholder="Ej: Collective Mining")
            mun_enc    = st.text_input("Municipios / territorios:", placeholder="Ej: Marmato y Supía")
            per_enc    = st.text_input("Período de evaluación:", placeholder="Ej: Primer semestre 2026")
        with c2:
            esc_enc    = st.radio("Escala Likert:", ["1 a 3","1 a 5"], horizontal=True)
            esc_max    = 3 if esc_enc == "1 a 3" else 5
            st.caption(f"🟢 ≥{esc_max*0.9:.1f} | 🟡 {esc_max*0.67:.1f}–{esc_max*0.9-0.01:.2f} | 🔴 <{esc_max*0.67:.1f}")
            proy_txt   = st.text_area("Lista de proyectos (uno por línea):", height=115,
                                      placeholder="Proyecto Acueducto\nProyecto Educación\n...")
            proy_lista = [p.strip() for p in proy_txt.split("\n") if p.strip()]

    # ── Carga de archivos ─────────────────────────────────────────
    st.subheader("📂 Carga de archivos")
    st.info("Puedes subir uno, dos o los tres archivos. El análisis se adapta a lo disponible.")
    cf1, cf2, cf3 = st.columns(3)
    with cf1:
        arc_com = st.file_uploader("👥 Comunidad",  type=["xlsx"], key="f_com")
    with cf2:
        arc_ali = st.file_uploader("🤝 Aliados",    type=["xlsx"], key="f_ali")
    with cf3:
        arc_emp = st.file_uploader("🏢 Empresa",    type=["xlsx"], key="f_emp")

    arcs = {"Comunidad": arc_com, "Aliados": arc_ali, "Empresa": arc_emp}
    arcs_act = {k: v for k, v in arcs.items() if v is not None}

    if arcs_act:
        dfs_ok = {}
        for gr, arc in arcs_act.items():
            try:
                dfs_ok[gr] = pd.read_excel(arc)
            except Exception as e_r:
                st.error(f"Error leyendo {gr}: {e_r}")

        if dfs_ok:
            cols_ref = list(dfs_ok[list(dfs_ok.keys())[0]].columns)

            # ── Configuración de columnas ─────────────────────────
            st.subheader("🗂️ Configuración de columnas")
            opc_proy  = ["(sin columna de proyecto)"] + cols_ref
            col_proy  = st.selectbox("Columna que identifica el proyecto:", opc_proy)
            col_proy  = None if col_proy.startswith("(sin") else col_proy
            cm1, cm2  = st.columns(2)
            with cm1:
                cols_lk  = st.multiselect("Columnas Likert (numéricas):", cols_ref)
                col_cal  = st.selectbox("Calificación 1-5 (si existe):",
                                        ["(ninguna)"] + cols_ref)
                col_cal  = None if col_cal == "(ninguna)" else col_cal
            with cm2:
                cols_sn  = st.multiselect("Columnas Sí / No:", cols_ref)
                cols_txt = st.multiselect("Columnas texto abierto:", cols_ref)
            cm3, cm4 = st.columns(2)
            with cm3:
                col_muni = st.selectbox("Columna municipio (opcional):",
                                        ["(ninguna)"] + cols_ref)
                col_muni = None if col_muni == "(ninguna)" else col_muni
                col_ver  = st.selectbox("Columna vereda (opcional):",
                                        ["(ninguna)"] + cols_ref)
                col_ver  = None if col_ver == "(ninguna)" else col_ver
            with cm4:
                col_part = st.selectbox(
                    "Columna participantes (si cada fila es UN participante, deja (ninguna)):",
                    ["(ninguna)"] + cols_ref)
                col_part = None if col_part == "(ninguna)" else col_part

            # ── Fase 1: Validación ────────────────────────────────
            st.subheader("📋 Fase 1 — Validación de datos")
            if st.button("🔍 Validar datos"):
                st.session_state["enc_fase1_ok"] = False
                resumen_val = []
                for gr_v, df_v in dfs_ok.items():
                    alertas_v = []
                    for cl_v in cols_lk:
                        if cl_v in df_v.columns:
                            col_num_v = pd.to_numeric(df_v[cl_v], errors="coerce")
                            no_num_v  = col_num_v.isna().sum() - df_v[cl_v].isna().sum()
                            if no_num_v > 0:
                                alertas_v.append(f"'{cl_v}': {no_num_v} no numéricos")
                            fuera_v = ((col_num_v < 1)|(col_num_v > esc_max)).sum()
                            if fuera_v > 0:
                                alertas_v.append(f"'{cl_v}': {fuera_v} fuera de 1-{esc_max}")
                    pdet = {}
                    if col_proy and col_proy in df_v.columns:
                        pdet = df_v[col_proy].value_counts().to_dict()
                        if gr_v == "Aliados":
                            posibles_id = [c for c in df_v.columns
                                           if any(x in c.lower()
                                                  for x in ["nombre","respondente","id","persona"])]
                            if posibles_id:
                                n_pers = df_v[posibles_id[0]].nunique()
                                st.info(f"ℹ️ **Aliados:** {n_pers} personas únicas, "
                                        f"{len(df_v)} evaluaciones (col: '{posibles_id[0]}')")
                        if proy_lista and pdet:
                            det_s_lower = set(normalizar_str(k) for k in pdet)
                            esp_s_lower = set(normalizar_str(p) for p in proy_lista)
                            det_orig = {normalizar_str(k): str(k).strip() for k in pdet}
                            esp_orig = {normalizar_str(p): p.strip() for p in proy_lista}
                            solo_datos = det_s_lower - esp_s_lower
                            solo_lista = esp_s_lower - det_s_lower
                            if solo_datos:
                                st.warning(f"**{gr_v}** — En datos pero NO en tu lista. "
                                           f"Copia exactamente: {', '.join(det_orig[k] for k in solo_datos)}")
                            if solo_lista:
                                st.warning(f"**{gr_v}** — En tu lista pero SIN datos. "
                                           f"Revisa mayúsculas/tildes: {', '.join(esp_orig[k] for k in solo_lista)}")
                        if pdet:
                            st.write(f"**{gr_v} — Proyectos detectados:**")
                            st.dataframe(pd.DataFrame(list(pdet.items()),
                                         columns=["Proyecto", f"Filas ({gr_v})"]),
                                         use_container_width=True)
                    resumen_val.append({
                        "Grupo": gr_v,
                        "Filas": len(df_v),
                        "Proyectos": len(pdet) if pdet else "N/A",
                        "Alertas": "; ".join(alertas_v) if alertas_v else "✅ Sin alertas"
                    })
                st.dataframe(pd.DataFrame(resumen_val), use_container_width=True)
                st.session_state["enc_fase1_ok"] = True
                st.success("✅ Validación completada. Presiona **▶ Analizar encuestas** para continuar.")

            # ── Fases 2-3-4: Análisis ─────────────────────────────
            if st.session_state.get("enc_fase1_ok"):
                contexto_enc = st.selectbox("Contexto del análisis:",
                    ["evaluación de inversión social y desarrollo comunitario",
                     "consultoría y diagnóstico organizacional",
                     "diagnóstico territorial participativo"])

                if st.button("▶ Analizar encuestas"):
                    # Determinar proyectos desde los datos reales (sin comparar strings)
                    todos_proy = set()
                    for gr_p, df_p in dfs_ok.items():
                        if col_proy and col_proy in df_p.columns:
                            todos_proy.update(df_p[col_proy].dropna().astype(str).str.strip().unique())
                    if not todos_proy and proy_lista:
                        todos_proy = set(proy_lista)
                    proyectos_a = sorted(todos_proy) if todos_proy else ["(análisis global)"]

                    filas_consolidado, filas_acuerdo = [], []
                    filas_sino, filas_cal, filas_tri = [], [], []
                    filas_cual, sin_clasificar_enc   = [], []

                    # ── Participación territorial (si el usuario configuró cols geo) ─
                    _col_muni = col_muni if "col_muni" in vars() else None
                    _col_ver  = col_ver  if "col_ver"  in vars() else None
                    _col_part = col_part if "col_part" in vars() else None

                    if _col_muni or _col_ver:
                        st.subheader("👥 Participación por territorio")
                        df_geo_list = []
                        for gr_geo, df_geo_g in dfs_ok.items():
                            df_tmp = df_geo_g.copy()
                            df_tmp["_grupo"] = gr_geo
                            if _col_part and _col_part in df_tmp.columns:
                                df_tmp["_part"] = pd.to_numeric(df_tmp[_col_part], errors="coerce").fillna(1)
                            else:
                                df_tmp["_part"] = 1
                            df_geo_list.append(df_tmp)
                        df_geo = pd.concat(df_geo_list, ignore_index=True)

                        if _col_muni and _col_muni in df_geo.columns:
                            pm_enc = (df_geo.groupby(_col_muni)["_part"]
                                      .sum().reset_index()
                                      .sort_values("_part", ascending=False))
                            pm_enc.columns = ["Municipio", "Participantes"]
                            col_pm1, col_pm2 = st.columns(2)
                            with col_pm1:
                                st.plotly_chart(
                                    px.bar(pm_enc, x="Municipio", y="Participantes",
                                           color="Municipio", text="Participantes",
                                           title="Participantes por municipio"),
                                    use_container_width=True)
                            with col_pm2:
                                st.dataframe(pm_enc, use_container_width=True)

                        if _col_ver and _col_ver in df_geo.columns:
                            group_cols = ([_col_muni, _col_ver]
                                          if _col_muni and _col_muni in df_geo.columns
                                          else [_col_ver])
                            pv_enc = (df_geo.groupby(group_cols)["_part"]
                                      .sum().reset_index())
                            if len(group_cols) == 2:
                                pv_enc.columns = ["Municipio", "Vereda", "Participantes"]
                                pv_enc = pv_enc.sort_values(["Municipio","Participantes"],
                                                             ascending=[True, False])
                                fig_pv = px.bar(pv_enc, x="Vereda", y="Participantes",
                                                color="Municipio", text="Participantes",
                                                title="Participantes por vereda")
                            else:
                                pv_enc.columns = ["Vereda", "Participantes"]
                                pv_enc = pv_enc.sort_values("Participantes", ascending=False)
                                fig_pv = px.bar(pv_enc, x="Vereda", y="Participantes",
                                                text="Participantes",
                                                title="Participantes por vereda")
                            st.plotly_chart(fig_pv, use_container_width=True)
                            st.dataframe(pv_enc, use_container_width=True)

                        pg_enc = (df_geo.groupby("_grupo")["_part"]
                                  .sum().reset_index())
                        pg_enc.columns = ["Grupo", "Participantes"]
                        total_enc = int(pg_enc["Participantes"].sum())
                        pg_enc["% del total"] = (pg_enc["Participantes"]/total_enc*100).round(1).astype(str)+"%"
                        st.metric("👥 Total participantes (todos los grupos)", total_enc)
                        st.dataframe(pg_enc, use_container_width=True)

                    # ── Fase 2: Likert — usando groupby (sin comparación de strings) ─
                    if cols_lk:
                        st.subheader("📊 Fase 2 — Análisis Likert por proyecto")
                        st.caption("🟢 Destacado | 🟡 Aceptable | 🔴 Crítico | ⚠️indiv. = n≤2")

                        for gr, df_g in dfs_ok.items():
                            # Convertir todas las columnas Likert a numérico de una sola vez
                            df_lk = df_g.copy()
                            for cl in cols_lk:
                                if cl in df_lk.columns:
                                    df_lk[cl] = df_lk[cl].apply(
                                        lambda x: texto_a_likert(x, esc_max))

                            if col_proy and col_proy in df_lk.columns:
                                # Agrupar por proyecto — pandas hace el trabajo
                                cols_lk_disp = [c for c in cols_lk if c in df_lk.columns]
                                grp = df_lk.groupby(col_proy)[cols_lk_disp]
                                promedios = grp.mean()
                                conteos   = grp.count()

                                st.write(f"**{gr}** — Promedios Likert por proyecto:")
                                # Tabla con semáforos
                                df_sem = promedios.copy()
                                for col_s in df_sem.columns:
                                    df_sem[col_s] = df_sem[col_s].apply(
                                        lambda v: semaforo_likert(v, esc_max)
                                        if pd.notna(v) else "⬜")
                                st.dataframe(df_sem, use_container_width=True)

                                # Tabla numérica limpia
                                st.write("Vista numérica:")
                                st.dataframe(promedios.round(2), use_container_width=True)

                                # % de acuerdo
                                st.write("% de acuerdo (respuestas con valor máximo):")
                                df_ac = df_lk.groupby(col_proy)[cols_lk_disp].apply(
                                    lambda g: (g == esc_max).mean() * 100).round(1)
                                st.dataframe(df_ac.applymap(lambda v: f"{v:.0f}%"
                                    if pd.notna(v) else "—"), use_container_width=True)

                                # Acumular para exportar y triangulación
                                for proy_g in promedios.index:
                                    fila_c  = {"Proyecto": proy_g, "Grupo": gr}
                                    fila_ac = {"Proyecto": proy_g, "Grupo": gr}
                                    for cl in cols_lk_disp:
                                        v   = promedios.loc[proy_g, cl]
                                        n   = int(conteos.loc[proy_g, cl])
                                        av  = " ⚠️indiv." if n <= 2 else ""
                                        fila_c[f"{cl} (n={n}{av})"]  = semaforo_likert(v, esc_max)
                                        fila_ac[f"{cl}"] = f"{(df_lk[df_lk[col_proy]==proy_g][cl]==esc_max).mean()*100:.0f}%"
                                    filas_consolidado.append(fila_c)
                                    filas_acuerdo.append(fila_ac)

                                # Heatmap
                                fig_hm = px.imshow(
                                    promedios.T,
                                    color_continuous_scale="RdYlGn",
                                    zmin=1, zmax=esc_max, text_auto=".2f",
                                    title=f"Mapa de calor — {gr}",
                                    aspect="auto",
                                    labels={"x": "Proyecto", "y": "Indicador", "color": "Promedio"})
                                st.plotly_chart(fig_hm, use_container_width=True)
                            else:
                                # Sin columna de proyecto: análisis global
                                cols_lk_disp = [c for c in cols_lk if c in df_lk.columns]
                                promedios_g = df_lk[cols_lk_disp].mean().round(2)
                                st.write(f"**{gr}** — Análisis global (sin columna de proyecto):")
                                st.dataframe(promedios_g.to_frame("Promedio"), use_container_width=True)

                    # ── Fase 2: Sí/No — usando groupby ────────────
                    if cols_sn:
                        st.subheader("✅ Preguntas Sí / No")
                        for gr, df_g in dfs_ok.items():
                            df_sn = df_g.copy()
                            for cs in cols_sn:
                                if cs in df_sn.columns:
                                    df_sn[cs] = df_sn[cs].apply(
                                        lambda x: 1 if str(x).strip().lower()
                                        in ["sí","si","s","1","true","yes"] else (
                                        0 if str(x).strip().lower()
                                        in ["no","n","0","false"] else np.nan))
                            cols_sn_disp = [c for c in cols_sn if c in df_sn.columns]
                            if col_proy and col_proy in df_sn.columns and cols_sn_disp:
                                pct_sn = (df_sn.groupby(col_proy)[cols_sn_disp]
                                         .mean() * 100).round(1)
                                st.write(f"**{gr}** — % de respuestas Sí:")
                                st.dataframe(pct_sn.applymap(lambda v: f"{v:.0f}%"
                                    if pd.notna(v) else "—"), use_container_width=True)
                                for proy_g in pct_sn.index:
                                    fila_sn = {"Proyecto": proy_g, "Grupo": gr}
                                    for cs in cols_sn_disp:
                                        fila_sn[cs] = f"{pct_sn.loc[proy_g,cs]:.0f}% Sí"
                                    filas_sino.append(fila_sn)

                    # ── Fase 2: Calificación 1-5 — usando groupby ──
                    if col_cal:
                        st.subheader("⭐ Calificación general 1-5")
                        for gr, df_g in dfs_ok.items():
                            if col_cal not in df_g.columns:
                                continue
                            df_cal = df_g.copy()
                            df_cal[col_cal] = pd.to_numeric(df_cal[col_cal], errors="coerce")
                            if col_proy and col_proy in df_cal.columns:
                                cal_grp = df_cal.groupby(col_proy)[col_cal].agg(["mean","count"])
                                cal_grp.columns = ["Promedio","n"]
                                cal_grp["Semáforo"] = cal_grp["Promedio"].apply(semaforo_rating)
                                st.write(f"**{gr}** — Calificación 1-5 por proyecto:")
                                st.dataframe(cal_grp, use_container_width=True)
                                for proy_g, row_c in cal_grp.iterrows():
                                    av = " ⚠️indiv." if row_c["n"] <= 2 else ""
                                    filas_cal.append({
                                        "Proyecto": proy_g, "Grupo": gr,
                                        f"Calificación (n={int(row_c['n'])}{av})": row_c["Semáforo"]})

                    # ── Fase 3: Triangulación — usando datos acumulados ─
                    if cols_lk and len(dfs_ok) >= 2 and filas_consolidado:
                        st.subheader("🔺 Fase 3 — Triangulación entre grupos")
                        umbral_t = 0.5 if esc_max == 3 else 1.0
                        st.caption(f"Convergencia = diferencia MAX-MIN ≤ {umbral_t}")

                        # Construir tabla pivote: proyecto × grupo → promedio por indicador
                        for cl in cols_lk:
                            # Agrupar promedios por proyecto y grupo
                            datos_tri = {}
                            for gr, df_g in dfs_ok.items():
                                if cl not in df_g.columns or not (col_proy and col_proy in df_g.columns):
                                    continue
                                df_t = df_g.copy()
                                df_t[cl] = df_t[cl].apply(lambda x: texto_a_likert(x, esc_max))
                                for proy_g, v in df_t.groupby(col_proy)[cl].mean().items():
                                    if pd.notna(v):
                                        datos_tri.setdefault(str(proy_g), {})[gr] = round(v, 2)

                            for proy_g, vals_t in datos_tri.items():
                                if len(vals_t) >= 2:
                                    rng_t = max(vals_t.values()) - min(vals_t.values())
                                    filas_tri.append({
                                        "Proyecto": proy_g, "Indicador": cl,
                                        "Comunidad": vals_t.get("Comunidad", "Sin dato"),
                                        "Aliados":   vals_t.get("Aliados",   "Sin dato"),
                                        "Empresa":   vals_t.get("Empresa",   "Sin dato"),
                                        "Rango MAX-MIN": round(rng_t, 2),
                                        "Resultado": convergencia(rng_t, esc_max)})

                        if filas_tri:
                            df_tri_enc = pd.DataFrame(filas_tri)
                            st.dataframe(df_tri_enc, use_container_width=True)
                            n_div = df_tri_enc["Resultado"].str.contains("Divergencia").sum()
                            if n_div > 0:
                                st.warning(f"⚠️ {n_div} divergencias detectadas.")
                        else:
                            st.info("Se necesitan al menos 2 grupos con datos en los mismos indicadores.")

                    # ── Fase 3: Cualitativo — clasificación en un solo paso ──
                    if cols_txt:
                        st.subheader("💬 Fase 3 — Análisis cualitativo")

                        # Paso 1: recopilar todos los textos con metadata
                        registros_cual = []  # [{idx, gr, ct, proy, texto}]
                        for gr, df_g in dfs_ok.items():
                            for ct in cols_txt:
                                if ct not in df_g.columns:
                                    continue
                                for idx_row, row_ct in df_g.iterrows():
                                    txt_ct = str(row_ct.get(ct, "")).strip()
                                    if not txt_ct or txt_ct == "nan" or len(txt_ct) < 3:
                                        continue
                                    proy_ct = (str(row_ct.get(col_proy, "N/A"))
                                               if col_proy and col_proy in df_g.columns
                                               else "N/A")
                                    registros_cual.append({
                                        "idx": len(registros_cual),
                                        "Grupo": gr, "Columna": ct,
                                        "Proyecto": proy_ct, "Texto": txt_ct})

                        # Paso 2: clasificar por keywords primero
                        temas_disponibles = list(KEYWORDS_TEMAS.keys())
                        temas_norm = {normalizar_str(t): t for t in temas_disponibles}

                        for reg in registros_cual:
                            reg["Tema"] = clasificar_texto_enc(reg["Texto"])[0]

                        # Paso 3: los que quedaron "Sin clasificar" → IA en lotes
                        sin_cl_idx = [r for r in registros_cual if r["Tema"] == "Sin clasificar"]

                        if sin_cl_idx:
                            temas_str_lk = "\n".join(f"{i+1}. {t}"
                                                      for i, t in enumerate(temas_disponibles))
                            with st.spinner(f"Clasificando {len(sin_cl_idx)} respuestas con IA..."):
                                for inicio in range(0, len(sin_cl_idx), 20):
                                    lote = sin_cl_idx[inicio:inicio + 15]
                                    textos_lote = "\n".join(
                                        f"{i+1}. {r['Texto'][:150]}"
                                        for i, r in enumerate(lote))
                                    prompt_cl = (
                                        "Eres un experto en análisis de inversión social comunitaria.\n"
                                        "Tienes esta lista de temas numerados:\n"
                                        + temas_str_lk + "\n\n"
                                        "Para cada respuesta, escribe SOLO el número del tema "
                                        "que mejor corresponda.\n"
                                        "IMPORTANTE: siempre elige el tema más cercano. "
                                        "Nunca escribas 0 — si la respuesta es corta o vaga, "
                                        "asigna el tema más probable en contexto de proyectos comunitarios.\n\n"
                                        "Respuestas:\n" + textos_lote + "\n\n"
                                        "Una línea por respuesta, solo el número:\n"
                                        "1. 3\n2. 7\n3. 12\nSolo el número, sin explicaciones.")
                                    try:
                                        resp_cl = client.chat.complete(
                                            model="mistral-small-latest",
                                            messages=[{"role": "user", "content": prompt_cl}],
                                            max_tokens=100)
                                        lineas_cl = resp_cl.choices[0].message.content.strip().split("\n")
                                        for i, reg in enumerate(lote):
                                            if i >= len(lineas_cl):
                                                break
                                            linea = lineas_cl[i].strip()
                                            # Extraer el número de la línea (ej: "1. 3" → "3")
                                            num_str = linea.split(".")[-1].strip()
                                            try:
                                                num = int(num_str)
                                                if 1 <= num <= len(temas_disponibles):
                                                    reg["Tema"] = temas_disponibles[num - 1] + " (IA)"
                                                else:
                                                    # 0 o fuera de rango: asignar "Mejora y sugerencias"
                                                    reg["Tema"] = "Mejora y sugerencias (IA)"
                                            except (ValueError, IndexError):
                                                # Respuesta no numérica: intentar parsear texto
                                                for t in temas_disponibles:
                                                    if normalizar_str(t) in normalizar_str(num_str):
                                                        reg["Tema"] = t + " (IA)"
                                                        break
                                                else:
                                                    reg["Tema"] = "Mejora y sugerencias (IA)"
                                    except Exception:
                                        pass  # lote falla → quedan Sin clasificar

                        # Paso 4: construir filas_cual y sin_clasificar_enc
                        for reg in registros_cual:
                            filas_cual.append({
                                "Grupo":    reg["Grupo"],
                                "Proyecto": reg["Proyecto"],
                                "Columna":  reg["Columna"],
                                "Tema":     reg["Tema"],
                                "Texto":    reg["Texto"][:200]})
                            if reg["Tema"] == "Sin clasificar":
                                sin_clasificar_enc.append({
                                    "Grupo":   reg["Grupo"],
                                    "Columna": reg["Columna"],
                                    "Texto":   reg["Texto"]})

                        # Paso 5: mostrar resultados
                        if filas_cual:
                            df_cual = pd.DataFrame(filas_cual)
                            clasificados   = df_cual[df_cual["Tema"] != "Sin clasificar"]
                            n_sin_cl       = (df_cual["Tema"] == "Sin clasificar").sum()
                            n_ia           = df_cual["Tema"].str.endswith("(IA)").sum()
                            total_txt      = len(df_cual)

                            st.caption(
                                f"Total respuestas: {total_txt} | "
                                f"Clasificadas por palabras clave: {total_txt - n_sin_cl - n_ia} | "
                                f"Clasificadas por IA: {n_ia} | "
                                f"Sin clasificar: {n_sin_cl}")

                            top_temas = (clasificados["Tema"].str.replace(" (IA)", "", regex=False)
                                        .value_counts().head(14).reset_index())
                            top_temas.columns = ["Tema", "Frecuencia"]
                            fig_t = px.bar(top_temas, x="Frecuencia", y="Tema",
                                          orientation="h",
                                          title="Temas más frecuentes en respuestas abiertas",
                                          color="Frecuencia", color_continuous_scale="Blues")
                            fig_t.update_layout(yaxis=dict(autorange="reversed"))
                            st.plotly_chart(fig_t, use_container_width=True)

                            if sin_clasificar_enc:
                                with st.expander(
                                        f"⚠️ {len(sin_clasificar_enc)} respuestas Sin clasificar "
                                        f"({len(sin_clasificar_enc)*100//total_txt}% del total)"):
                                    st.dataframe(pd.DataFrame(sin_clasificar_enc),
                                                 use_container_width=True)
                                    st.caption(
                                        "Estas respuestas son muy cortas, muy genéricas o "
                                        "mencionan temas fuera del diccionario. "
                                        "Puedes reclasificarlas manualmente en la hoja "
                                        "5_Abiertas del Excel descargado.")

                    # ── Fase 4: Informe IA ────────────────────────
                    st.subheader("📄 Fase 4 — Informe ejecutivo IA")
                    resumen_ia = (f"Organización: {org_enc} | Cliente: {cli_enc}\n"
                                  f"Municipios: {mun_enc} | Período: {per_enc}\n"
                                  f"Grupos: {', '.join(dfs_ok.keys())}\n"
                                  f"Proyectos: {', '.join(str(p) for p in proyectos_a[:15])}\n")
                    if filas_consolidado:
                        resumen_ia += f"\nPromedios Likert:\n{pd.DataFrame(filas_consolidado).to_string(index=False)}\n"
                    if filas_tri:
                        n_div_ia = sum(1 for f in filas_tri if "Divergencia" in str(f.get("Resultado","")))
                        resumen_ia += f"\nTriangulación: {len(filas_tri)} cruces, {n_div_ia} divergencias.\n"
                    if filas_cual:
                        top_t_ia = (pd.DataFrame(filas_cual)[pd.DataFrame(filas_cual)["Tema"]!="Sin clasificar"]
                                    ["Tema"].value_counts().head(5).index.tolist())
                        resumen_ia += f"\nTemas principales: {', '.join(top_t_ia)}\n"

                    with st.spinner("Generando informe con IA..."):
                        informe_enc = generar_informe_ia(resumen_ia, contexto_enc)
                        st.session_state.informe_enc = informe_enc
                    st.markdown(informe_enc)

                    # ── Descarga Excel con fórmulas ───────────────
                    st.subheader("⬇ Descargas")
                    cfg_dict = {
                        "organizacion": org_enc, "cliente": cli_enc,
                        "municipios": mun_enc,   "periodo": per_enc,
                        "escala_max": esc_max,   "col_proyecto": col_proy,
                        "cols_likert": cols_lk,  "cols_sino": cols_sn,
                        "col_cal": col_cal,      "cols_texto": cols_txt,
                        "proyectos_lista": proyectos_a,
                    }
                    res_dict = {
                        "filas_consolidado": filas_consolidado,
                        "filas_tri": filas_tri,
                        "filas_cual": filas_cual,
                        "sin_clasificar": sin_clasificar_enc,
                    }

                    with st.spinner("Construyendo Excel con fórmulas..."):
                        wb_enc = crear_excel_encuesta_formulas(dfs_ok, cfg_dict, res_dict)
                        buf_enc = io.BytesIO()
                        wb_enc.save(buf_enc)
                        buf_enc.seek(0)

                    cd1, cd2 = st.columns(2)
                    with cd1:
                        st.download_button(
                            "⬇ Descargar Excel con fórmulas",
                            data=buf_enc,
                            file_name="analisis_encuestas.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="dl_enc_xlsx")
                    with cd2:
                        if st.session_state.get("informe_enc"):
                            st.download_button(
                                "⬇ Descargar informe TXT",
                                data=st.session_state.informe_enc.encode("utf-8"),
                                file_name="informe_encuestas.txt",
                                key="dl_enc_txt")
    else:
        st.info("👆 Sube al menos un archivo de encuesta para comenzar.")


# ════════════════════════════════════════════════════════════════
# MÓDULO 2 — CARTOGRAFÍA SOCIAL
# ════════════════════════════════════════════════════════════════
elif modo == "🗺️ Cartografía Social":
    st.header("Módulo de Cartografía Social")
    st.write("Sube tu Excel con columnas: año, semestre, municipio, vereda, "
             "participantes, componentes y lineas de inversion.")

    archivo = st.file_uploader("Sube tu archivo Excel", type=["xlsx"])

    if archivo:
        archivo_hash = md5_archivo(archivo)
        df = pd.read_excel(archivo)
        df.columns = [c.strip().lower() for c in df.columns]
        st.subheader("Vista previa de tus datos")
        st.dataframe(df.head())

        columnas = list(df.columns)
        fila_config = df.iloc[0]
        df = df.iloc[1:].reset_index(drop=True)

        cols_meta      = [c for c in columnas if str(fila_config[c]).strip().lower()=="meta"]
        col_lineas_lst = [c for c in columnas if str(fila_config[c]).strip().lower()=="lineas"]
        col_lineas     = col_lineas_lst[0] if col_lineas_lst else "lineas de inversion"
        cols_comp_def  = [c for c in columnas if str(fila_config[c]).strip().lower()=="componente"]
        cols_componentes = st.multiselect("Selecciona los componentes a analizar:",
                                          cols_comp_def, default=cols_comp_def)
        contexto = st.selectbox("Contexto del análisis:",
            ["consultoría y diagnóstico organizacional",
             "bienestar social y desarrollo comunitario",
             "diagnóstico territorial participativo"])

        cache_key = f"{archivo_hash}_{','.join(sorted(cols_componentes))}_{contexto}"

        if cols_componentes and st.button("▶ Analizar"):
            # ── NIVEL 1: Caché en sesión (instantáneo) ────────────
            en_sesion = (st.session_state.get("cache_key_saved") == cache_key and
                         st.session_state.get("cache_result_cart") is not None)

            # ── NIVEL 2: Caché en GitHub (entre sesiones) ─────────
            en_github = False
            if not en_sesion:
                with st.spinner("Buscando análisis guardado en GitHub..."):
                    datos_gh = leer_cache_github(archivo_hash)
                if datos_gh is not None:
                    try:
                        df_result = cache_a_df(datos_gh["df_result"])
                        st.session_state.cache_result_cart = df_result
                        st.session_state.cache_key_saved   = cache_key
                        st.session_state.cache_cat_cart    = datos_gh["cat_comp"]
                        st.session_state.cache_cat_vereda  = datos_gh["cat_ver"]
                        en_github = True
                    except Exception:
                        en_github = False

            if en_sesion:
                st.info("✅ Mismos datos (sesión activa) — resultados instantáneos.")
                df_result = st.session_state.cache_result_cart
                total_participantes = (pd.to_numeric(df["participantes"],errors="coerce").sum()
                                       if "participantes" in df.columns else 0)
            elif en_github:
                st.info("✅ Análisis encontrado en GitHub — cargando resultados guardados.")
                total_participantes = (pd.to_numeric(df["participantes"],errors="coerce").sum()
                                       if "participantes" in df.columns else 0)
            else:
                total_participantes = 0
                if "participantes" in df.columns:
                    total_participantes = pd.to_numeric(df["participantes"],errors="coerce").sum()
                    st.metric("👥 Total participantes", int(total_participantes))

                if "municipio" in df.columns and "vereda" in df.columns:
                    st.subheader("🗺️ Participación territorial")
                    pm = df.groupby("municipio")["participantes"].apply(
                        lambda x: pd.to_numeric(x,errors="coerce").sum()).reset_index()
                    pm.columns = ["Municipio","Participantes"]
                    st.plotly_chart(px.bar(pm,x="Municipio",y="Participantes",
                                          color="Municipio",title="Participantes por municipio"),
                                   use_container_width=True)
                    pv = df.groupby(["municipio","vereda"])["participantes"].apply(
                        lambda x: pd.to_numeric(x,errors="coerce").sum()).reset_index()
                    pv.columns = ["Municipio","Vereda","Participantes"]
                    st.plotly_chart(px.bar(pv,x="Vereda",y="Participantes",
                                          color="Municipio",title="Participantes por vereda"),
                                   use_container_width=True)
                    if "año" in df.columns:
                        df["año"] = df["año"].astype(str).str.replace(".0","",regex=False)
                        pa = df.groupby("año")["participantes"].apply(
                            lambda x: pd.to_numeric(x,errors="coerce").sum()).reset_index()
                        pa.columns = ["Año","Participantes"]
                        st.plotly_chart(px.bar(pa,x="Año",y="Participantes",
                                              color="Año",title="Participantes por año",
                                              category_orders={"Año":sorted(pa["Año"].unique())}),
                                       use_container_width=True)
                    if "semestre" in df.columns:
                        ps = df.groupby("semestre")["participantes"].apply(
                            lambda x: pd.to_numeric(x,errors="coerce").sum()).reset_index()
                        ps.columns = ["Semestre","Participantes"]
                        st.plotly_chart(px.bar(ps,x="Semestre",y="Participantes",
                                              color="Semestre",title="Participantes por semestre"),
                                       use_container_width=True)

                with st.spinner("Fragmentando frases y procesando con el modelo semántico..."):
                    registros = []
                    for _, fila in df.iterrows():
                        lineas_c = (str(fila.get(col_lineas,"")).split(",")
                                    if col_lineas in df.columns else [])
                        lineas_c = [l.strip() for l in lineas_c if l.strip()]
                        for comp in cols_componentes:
                            celda = str(fila.get(comp,""))
                            if celda and celda != "nan":
                                frases = [f.strip() for f in celda.split(".") if len(f.strip())>5]
                                for frase in frases:
                                    reg = {"componente": comp, "frase": frase,
                                           "lineas_disponibles": lineas_c}
                                    for cm in cols_meta:
                                        reg[cm] = fila.get(cm,"")
                                    registros.append(reg)

                    df_frases = pd.DataFrame(registros)
                    st.write(f"Total de frases extraídas: {len(df_frases)}")
                    textos_l = df_frases["frase"].tolist()
                    vectores = np.array(modelo.encode(textos_l,show_progress_bar=False,batch_size=16))

                    resultados = []
                    for comp in cols_componentes:
                        mask = df_frases["componente"]==comp
                        vc = vectores[mask]
                        sub = df_frases[mask].reset_index(drop=True)
                        n_opt = (encontrar_clusters_optimos(vc,max_k=min(6,len(sub)-1))
                                 if len(sub)>=4 else min(2,len(sub)))
                        if n_opt>=2:
                            km = KMeans(n_clusters=n_opt,random_state=42,n_init=10)
                            clusters = km.fit_predict(vc)
                            pesos = cosine_similarity(vc,km.cluster_centers_).max(axis=1).round(3)
                        else:
                            clusters = [0]*len(sub)
                            pesos = [1.0]*len(sub)
                        for i, (_, row_r) in enumerate(sub.iterrows()):
                            res = {"componente":comp,"frase":row_r["frase"],
                                   "grupo_num":clusters[i],"peso_semantico":pesos[i],
                                   "lineas_disponibles":row_r["lineas_disponibles"]}
                            for cm in cols_meta:
                                res[cm] = row_r.get(cm,"")
                            resultados.append(res)

                    df_result = pd.DataFrame(resultados)

                letras = "ABCDEFGHIJKLMNÑOPQRSTUVWXYZ"
                for comp in cols_componentes:
                    mask = df_result["componente"]==comp
                    gu = sorted(df_result[mask]["grupo_num"].unique())
                    mapa = {g: f"Grupo {letras[i]}" for i,g in enumerate(gu)}
                    df_result.loc[mask,"grupo"] = df_result.loc[mask,"grupo_num"].map(mapa)

                with st.spinner("Asociando frases a líneas de inversión con IA..."):
                    ncomp = ["economico","social","ambiental","gobernanza","alianzas",
                             "proyeccion","governance","económico","proyección"]
                    for comp in cols_componentes:
                        mask = df_result["componente"]==comp
                        sub  = df_result[mask].reset_index(drop=True)
                        if len(sub)==0:
                            continue
                        lin_disp = sub.iloc[0]["lineas_disponibles"]
                        if not lin_disp:
                            df_result.loc[mask,"lineas_inversion"] = "Sin líneas definidas"
                            continue
                        frases_l = sub["frase"].tolist()
                        lin_str  = "\n".join(f"- {l}" for l in lin_disp)
                        for ini in range(0,len(frases_l),20):
                            sublote = frases_l[ini:ini+20]
                            frases_str = "\n".join(f"{i+1}. {f}" for i,f in enumerate(sublote))
                            prompt_l = (f"Eres un experto en desarrollo territorial (IAP, Fals Borda, Freire).\n"
                                        f"Líneas disponibles (usa EXACTAMENTE estos nombres):\n{lin_str}\n\n"
                                        f"Para cada frase, indica las líneas correspondientes.\nFrases:\n{frases_str}\n\n"
                                        f"Formato: 1. Linea1, Linea2\n2. Linea3\nSin explicaciones.")
                            try:
                                resp_l = client.chat.complete(
                                    model="mistral-small-latest",
                                    messages=[{"role":"user","content":prompt_l}],
                                    max_tokens=1500)
                                lineas_r = resp_l.choices[0].message.content.strip().split("\n")
                                for i,frase in enumerate(sublote):
                                    if i<len(lineas_r):
                                        lin_raw = lineas_r[i].split(". ",1)[-1].strip().rstrip(".")
                                        partes  = [p.strip().rstrip(".") for p in lin_raw.split(",")]
                                        norm_ps = []
                                        for parte in partes:
                                            pl = parte.lower().strip()
                                            if any(nc in pl for nc in ncomp):
                                                continue
                                            mejor = min(lin_disp,
                                                        key=lambda l: sum(c not in l.lower() for c in pl))
                                            if len(set(pl.split())&set(mejor.lower().split()))>=1:
                                                norm_ps.append(mejor)
                                        resultado_l = ", ".join(norm_ps) if norm_ps else lin_disp[0]
                                        idx_g = ini+i
                                        if idx_g<len(sub):
                                            df_result.loc[sub.index[idx_g],"lineas_inversion"] = resultado_l
                            except:
                                for i in range(len(sublote)):
                                    idx_g = ini+i
                                    if idx_g<len(sub):
                                        df_result.loc[sub.index[idx_g],"lineas_inversion"] = (
                                            lin_disp[0] if lin_disp else "No determinado")

                st.session_state.cache_result_cart = df_result
                st.session_state.cache_key_saved   = cache_key

                # ── NIVEL 3: Guardar en GitHub para futuras sesiones ──
                # (se hace después del análisis completo, incluyendo categorización)
                # Se ejecuta al final — ver bloque de guardado post-categorización

            # ── Filtros ───────────────────────────────────────────
            st.subheader("🔎 Filtros")
            ff1, ff2 = st.columns(2)
            with ff1:
                muni_f = st.multiselect("Filtrar por municipio:",
                    df_result["municipio"].dropna().unique() if "municipio" in df_result.columns else [],
                    default=list(df_result["municipio"].dropna().unique())
                    if "municipio" in df_result.columns else [])
            with ff2:
                comp_f = st.multiselect("Filtrar por componente:",
                                        cols_componentes, default=cols_componentes)

            df_fil = df_result.copy()
            if muni_f and "municipio" in df_fil.columns:
                df_fil = df_fil[df_fil["municipio"].isin(muni_f)]
            if comp_f:
                df_fil = df_fil[df_fil["componente"].isin(comp_f)]

            # ── Distribución por componente ───────────────────────
            st.subheader("📊 Distribución por componente")
            cnt_comp = df_fil["componente"].value_counts().reset_index()
            cnt_comp.columns = ["Componente","Frases"]
            st.plotly_chart(px.pie(cnt_comp,names="Componente",values="Frases",
                                   title="Distribución de frases por componente",hole=0.3),
                           use_container_width=True)

            # ── Líneas de inversión ───────────────────────────────
            st.subheader("📈 Líneas de inversión")
            lin_cnt = []
            for ls in df_fil["lineas_inversion"]:
                for l in str(ls).split(","):
                    nm = normalizar_linea(l.strip().rstrip(".").strip())
                    if nm:
                        lin_cnt.append(nm)
            if lin_cnt:
                cnt_lin = pd.DataFrame({"linea":lin_cnt})["linea"].value_counts().reset_index()
                cnt_lin.columns = ["Línea de inversión","Frecuencia"]
                st.plotly_chart(px.bar(cnt_lin,x="Línea de inversión",y="Frecuencia",
                                      color="Línea de inversión",
                                      title="Distribución por línea de inversión"),
                               use_container_width=True)

            # ── Cruce componente x línea ──────────────────────────
            st.subheader("🔀 Cruce componente × línea")
            cruce_v = []
            for _, row_cv in df_fil.iterrows():
                for l in str(row_cv["lineas_inversion"]).split(","):
                    nm = normalizar_linea(l.strip().rstrip(".").strip())
                    if nm:
                        cruce_v.append({"Componente":row_cv["componente"],"Línea":nm})
            if cruce_v:
                piv_cv = pd.DataFrame(cruce_v).groupby(["Componente","Línea"]).size().reset_index(name="n")
                st.plotly_chart(px.bar(piv_cv,x="Componente",y="n",color="Línea",
                                      barmode="group",title="Componentes por línea de inversión"),
                               use_container_width=True)

            # ── Grupos semánticos ─────────────────────────────────
            st.subheader("Grupos semánticos por componente")
            cols_g = st.columns(2)
            for i_g, comp_g in enumerate(comp_f):
                sub_g = df_fil[df_fil["componente"]==comp_g]
                cnt_g = sub_g["grupo"].value_counts().reset_index()
                cnt_g.columns = ["Grupo","Frecuencia"]
                fr_rep = {}
                for grp in cnt_g["Grupo"]:
                    top1 = sub_g[sub_g["grupo"]==grp].nlargest(1,"peso_semantico")
                    if len(top1)>0:
                        frt = top1.iloc[0]["frase"]
                        fr_rep[grp] = frt[:90]+"..." if len(frt)>90 else frt
                cnt_g["Tema"] = cnt_g["Grupo"].map(fr_rep)
                with cols_g[i_g%2]:
                    fig_g = px.bar(cnt_g,x="Grupo",y="Frecuencia",color="Grupo",
                                  title=f"{comp_g.upper()}",text="Frecuencia",
                                  hover_data={"Tema":True})
                    fig_g.update_traces(textposition="outside")
                    fig_g.update_layout(showlegend=False,height=350)
                    st.plotly_chart(fig_g,use_container_width=True)
                    for _, rg in cnt_g.iterrows():
                        st.caption(f"**{rg['Grupo']}:** {rg['Tema']}")

            # ── Cohesión semántica ────────────────────────────────
            st.subheader("Cohesión semántica por componente")
            pesos_c = df_fil.groupby("componente")["peso_semantico"].mean().round(3).reset_index()
            st.plotly_chart(px.bar(pesos_c,x="componente",y="peso_semantico",
                                  color="peso_semantico",color_continuous_scale="Teal",
                                  title="Cohesión semántica promedio"),
                           use_container_width=True)

            # ── Categorización ────────────────────────────────────
            if ("cache_cat_cart" not in st.session_state or
                    st.session_state.get("cache_key_saved")!=cache_key):
                cat_comp = {}
                cat_ver  = {}
                veredas_u = (sorted(df_fil["vereda"].dropna().unique())
                             if "vereda" in df_fil.columns else [])
                for comp_c in comp_f:
                    frases_c = df_fil[df_fil["componente"]==comp_c]["frase"].tolist()
                    cat_comp[comp_c] = categorizar_hallazgos(frases_c, contexto)
                    cat_ver[comp_c]  = {}
                    for ver in veredas_u:
                        fr_v = df_fil[(df_fil["componente"]==comp_c)&
                                      (df_fil["vereda"]==ver)]["frase"].tolist()
                        if fr_v:
                            cat_ver[comp_c][ver] = categorizar_hallazgos(fr_v, contexto)
                st.session_state.cache_cat_cart   = cat_comp
                st.session_state.cache_cat_vereda  = cat_ver
                # Guardar análisis completo en GitHub
                with st.spinner("Guardando análisis en GitHub para futuras sesiones..."):
                    datos_a_guardar = {
                        "df_result": df_a_cache(df_result),
                        "cat_comp":  cat_comp,
                        "cat_ver":   cat_ver,
                    }
                    ok_gh = escribir_cache_github(archivo_hash, datos_a_guardar)
                    if ok_gh:
                        st.success("✅ Análisis guardado en GitHub — la próxima sesión cargará al instante.")
                    else:
                        st.caption("ℹ️ No se pudo guardar en GitHub (continúa normalmente).")
            else:
                cat_comp = st.session_state.cache_cat_cart
                cat_ver  = st.session_state.cache_cat_vereda

            st.subheader("🗂️ Categorización por componente")
            for comp_c2 in comp_f:
                with st.expander(f"📌 {comp_c2}"):
                    st.markdown(cat_comp.get(comp_c2,"No disponible"))

            st.subheader("🗂️ Categorización por vereda")
            veredas_u2 = (sorted(df_fil["vereda"].dropna().unique())
                          if "vereda" in df_fil.columns else [])
            if veredas_u2:
                ver_sel = st.selectbox("Selecciona una vereda:", veredas_u2)
                for comp_c3 in comp_f:
                    if ver_sel in cat_ver.get(comp_c3,{}):
                        with st.expander(f"📌 {comp_c3}"):
                            st.markdown(cat_ver[comp_c3][ver_sel])

            # ── Resumen ejecutivo ─────────────────────────────────
            st.subheader("📋 Resumen ejecutivo")
            resumen_exec = df_fil.groupby("componente").agg(
                Total_frases=("frase","count"),
                Cohesion=("peso_semantico","mean")).round(3).reset_index()
            resumen_exec["Porcentaje"] = (
                resumen_exec["Total_frases"]/resumen_exec["Total_frases"].sum()*100
            ).round(1).astype(str) + "%"
            resumen_exec.columns = ["Componente","Total frases","Cohesión semántica","Porcentaje"]
            st.dataframe(resumen_exec, use_container_width=True)

            # ── Frases representativas ────────────────────────────
            st.subheader("💬 Frases más representativas")
            for comp_fr2 in comp_f:
                with st.expander(f"📌 {comp_fr2}"):
                    sub_fr2 = df_fil[df_fil["componente"]==comp_fr2]
                    top_fr2 = sub_fr2.nlargest(5,"peso_semantico")
                    fv2, vv2 = [], []
                    for _, row_fr2 in top_fr2.iterrows():
                        frase2 = row_fr2["frase"].strip()
                        if vv2:
                            sims2 = cosine_similarity(modelo.encode([frase2]),vv2)[0]
                            if any(s>0.75 for s in sims2):
                                continue
                        peso2 = round(float(row_fr2["peso_semantico"]),3)
                        rel2  = ("Alta relevancia 🔴" if peso2>=0.85
                                 else ("Media relevancia 🟡" if peso2>=0.65 else "Baja relevancia 🟢"))
                        grp2  = row_fr2["grupo"]
                        sub_g2 = sub_fr2[sub_fr2["grupo"]==grp2]
                        st.write(f"• {frase2} (peso: {peso2} — {rel2})")
                        if sub_g2.shape[0]>1:
                            st.caption(f"  💬 {sub_g2.shape[0]} frases similares en este grupo")
                            for otra in sub_g2[sub_g2["frase"]!=frase2]["frase"].head(2):
                                st.caption(f"  ↳ {otra}")
                        lin2 = str(row_fr2["lineas_inversion"])
                        if lin2 not in ["Sin líneas definidas","No determinado","nan"]:
                            st.caption(f"  Líneas: {lin2}")
                        fv2.append(frase2)
                        vv2.append(modelo.encode([frase2])[0])

            # ── Informe IA ────────────────────────────────────────
            st.subheader("📄 Informe ejecutivo IA")
            res_ia = (f"\nContexto: {contexto}\n"
                      f"Total participantes: {int(total_participantes)}\n"
                      f"Total frases: {len(df_fil)}\n"
                      f"Municipios: {', '.join(muni_f)}\n"
                      f"Componentes: {', '.join(comp_f)}\n"
                      f"Distribución:\n{cnt_comp.to_string(index=False)}\n"
                      f"Cohesión:\n{pesos_c.to_string(index=False)}\n"
                      "Frases representativas:\n")
            for comp_ia in comp_f:
                top_ia = df_fil[df_fil["componente"]==comp_ia].nlargest(2,"peso_semantico")
                res_ia += f"\n{comp_ia}:\n"
                for _, row_ia in top_ia.iterrows():
                    res_ia += f"  - {row_ia['frase']} (Líneas: {row_ia['lineas_inversion']})\n"

            with st.spinner("Generando informe ejecutivo con IA..."):
                informe_cart = generar_informe_ia(res_ia, contexto)
                st.session_state.informe_cart = informe_cart
            st.markdown(informe_cart)

            # ── Tabla detallada ───────────────────────────────────
            st.subheader("Tabla detallada")
            comp_vista = st.selectbox("Ver frases del componente:", comp_f)
            cols_v = [c for c in ["municipio","vereda","componente","frase","grupo",
                                   "peso_semantico","lineas_inversion"] if c in df_fil.columns]
            st.dataframe(df_fil[df_fil["componente"]==comp_vista][cols_v],
                         use_container_width=True)

            st.session_state.resultados_cart = df_fil

            # ── Descarga Excel con fórmulas ───────────────────────
            st.subheader("⬇ Descargas")
            dc1, dc2 = st.columns(2)
            with dc1:
                with st.spinner("Construyendo Excel con fórmulas y semáforos..."):
                    wb_cart = crear_excel_cartografia_formulas(
                        df_fil, comp_f, cat_comp, cat_ver,
                        cnt_comp, pesos_c, resumen_exec, contexto)
                    buf_cart = io.BytesIO()
                    wb_cart.save(buf_cart)
                    buf_cart.seek(0)
                st.download_button(
                    "⬇ Descargar análisis Excel (con fórmulas)",
                    data=buf_cart,
                    file_name="analisis_cartografia.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_cart_xlsx")
            with dc2:
                if st.session_state.informe_cart:
                    st.download_button(
                        "⬇ Descargar informe TXT",
                        data=st.session_state.informe_cart.encode("utf-8"),
                        file_name="informe_cartografia.txt",
                        key="dl_cart_txt")
